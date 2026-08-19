"""Compare GPT, Qwen 3B, and Qwen 7B for proper_term mode across dev_v1 term-list variants.

Writes one styled .xlsx file (default:
``experiments/01_term_expansion_by_model/report/dev_v1_proper_term_across_models.xlsx``)
with 9 data rows: 3 term-list variants (original, expand, cleaned), each with
3 language rows (ende, enru, enes). Only ``proper_term`` mode is included.
Reads ``metrics_summary.json`` from ``<variant_dir>/{gpt,qwen_3b,qwen_7b}/``.

Each value cell is colored by ranking that (model, language) combination's
value **across the 3 variants** (not across models), using the shared
Good/Bad/Neutral convention (see ``src/analysis/excel_style.py``): green =
best variant, red = worst variant, the strictly-middle variant is left
unfilled. This is a different axis than the ``best`` column, which colors
green/yellow by the best *model* within a single row.

Note: ``results/dev_v1/original/`` has no ``gpt``/``qwen_3b``/``qwen_7b``
subfolders directly — it's nested under ``zero_shot/`` or
``few_shot/`` (see ``report/README.md`` §3.4.1). This script defaults
to ``few_shot``; override with ``--original`` if you want the
``zero_shot`` variant instead.

Usage::

    python experiments/01_term_expansion_by_model/scripts/compare_proper_term_across_models_to_excel.py
    python experiments/01_term_expansion_by_model/scripts/compare_proper_term_across_models_to_excel.py --original results/dev_v1/original/zero_shot
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    best_label,
    label_fill,
    rank_fills,
)

LANG_ORDER = ("ende", "enru", "enes")
VARIANT_ORDER = ("original", "expand", "cleaned")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
BASELINE_LABELS = {
    "gpt": "GPT",
    "qwen_3b": "Qwen 3B",
    "qwen_7b": "Qwen 7B",
}
MODE = "proper_term"

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_proper_term_metrics(summary: dict[str, Any]) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(MODE)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {column: extract_metric(metrics, spec) for column, spec, _ in METRICS}
    return by_lang


def build_comparison(variant_dirs: dict[str, Path]) -> dict[tuple[str, str], dict[str, object]]:
    summaries: dict[tuple[str, str], dict[str, dict[str, float | None]]] = {}

    for variant, variant_dir in variant_dirs.items():
        for baseline_dir in BASELINE_DIRS:
            summary_path = variant_dir / baseline_dir / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(f"Missing metrics file: {summary_path}")
            summaries[(variant, baseline_dir)] = extract_proper_term_metrics(load_summary(summary_path))

    rows: dict[tuple[str, str], dict[str, object]] = {}
    for variant in VARIANT_ORDER:
        for lang in LANG_ORDER:
            baseline_metrics = {
                baseline_dir: summaries[(variant, baseline_dir)].get(lang) for baseline_dir in BASELINE_DIRS
            }
            if all(metrics is None for metrics in baseline_metrics.values()):
                continue

            row: dict[str, object] = {"data": variant, "language": lang}

            for column, _, _ in METRICS:
                labeled_values: dict[str, float | None] = {}
                for baseline_dir in BASELINE_DIRS:
                    metrics = baseline_metrics[baseline_dir]
                    value = metrics.get(column) if metrics else None
                    row[f"{baseline_dir}_{column}"] = value
                    labeled_values[BASELINE_LABELS[baseline_dir]] = value

                row[f"best_{column}"] = best_label(labeled_values)

            rows[(variant, lang)] = row

    return rows


def variant_rank_fill(
    rows: dict[tuple[str, str], dict[str, object]], lang: str, key: str, variant: str
) -> tuple[PatternFill, Font] | None:
    """Rank a (model, language) value against its counterpart in the other 2 variants."""
    values = {
        v: rows[(v, lang)][key]
        for v in VARIANT_ORDER
        if (v, lang) in rows and rows[(v, lang)][key] is not None
    }
    return rank_fills(values).get(variant)


def write_styled_excel(rows: dict[tuple[str, str], dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "proper_term"

    fixed_headers = ("data", "language")
    value_subheaders = tuple(BASELINE_LABELS[baseline_dir] for baseline_dir in BASELINE_DIRS) + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    row_offset = 3
    for variant in VARIANT_ORDER:
        for lang_idx, lang in enumerate(LANG_ORDER):
            record = rows.get((variant, lang))
            if record is None:
                continue

            data_cell = ws.cell(row=row_offset, column=1, value=variant if lang_idx == 0 else None)
            apply_cell_style(data_cell)

            lang_cell = ws.cell(row=row_offset, column=2, value=lang)
            apply_cell_style(lang_cell)

            for metric_idx, (column, _, _) in enumerate(METRICS):
                start_col = metric_start_col + metric_idx * cols_per_metric
                baseline_key_pairs = [(f"{baseline_dir}_{column}", baseline_dir) for baseline_dir in BASELINE_DIRS]

                for offset, (key, _baseline_dir) in enumerate(baseline_key_pairs):
                    cell = ws.cell(row=row_offset, column=start_col + offset, value=record[key])
                    fill_font = variant_rank_fill(rows, lang, key, variant)
                    apply_cell_style(
                        cell,
                        fill=fill_font[0] if fill_font else None,
                        font=fill_font[1] if fill_font else None,
                    )

                best_cell = ws.cell(
                    row=row_offset,
                    column=start_col + cols_per_metric - 1,
                    value=record[f"best_{column}"],
                )
                best_fill_font = label_fill(record[f"best_{column}"])
                apply_cell_style(
                    best_cell,
                    fill=best_fill_font[0] if best_fill_font else None,
                    font=best_fill_font[1] if best_fill_font else None,
                )

            row_offset += 1

    for variant_idx in range(len(VARIANT_ORDER)):
        start_row = 3 + variant_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    autofit_columns(ws)

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1 result folders",
    )
    parser.add_argument(
        "--original",
        type=Path,
        default=None,
        help=(
            "Path to the dev_v1/original results (default: "
            "<results-root>/dev_v1/original/few_shot — "
            "override with <results-root>/dev_v1/original/zero_shot if preferred)"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "experiments/01_term_expansion_by_model/report/dev_v1_proper_term_across_models.xlsx"
        ),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    variant_dirs = {
        "original": (args.original.resolve() if args.original else results_root / "dev_v1" / "original" / "few_shot"),
        "expand": results_root / "dev_v1" / "expand",
        "cleaned": results_root / "dev_v1" / "cleaned",
    }

    rows = build_comparison(variant_dirs)
    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_styled_excel(rows, output_path)

    print(
        f"Wrote {len(rows)} rows "
        f"({len(VARIANT_ORDER)} variants x {len(LANG_ORDER)} languages) to {output_path}"
    )


if __name__ == "__main__":
    main()
