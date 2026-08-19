"""Compare ende, enru, and enes language pairs across all baseline models and modes.

Writes one styled .xlsx file (default:
``experiments/02_term_expansion_by_language_pair/report/language_comparison.xlsx``)
with one sheet per dataset variant (``dev_v1_original_zero_shot``,
``dev_v1_original_few_shot``, ``dev_v1_expand``, ``dev_v1_cleaned``, ``dev_v2``), each
with 9 data rows: 3 modes (no_term, proper_term, random_term), each with 3 model rows
(GPT, Qwen 3B, Qwen 7B). The mode column is merged per block. Reads
``metrics_summary.json`` from ``<results-root>/<variant>/{gpt,qwen_3b,qwen_7b}/``.

Usage::

    python experiments/02_term_expansion_by_language_pair/scripts/compare_languages_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    best_label,
    label_fill,
)

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
BASELINE_LABELS = {
    "gpt": "GPT",
    "qwen_3b": "Qwen 3B",
    "qwen_7b": "Qwen 7B",
}

DATASET_VARIANTS = (
    ("dev_v1_original_zero_shot", Path("dev_v1/original/zero_shot")),
    ("dev_v1_original_few_shot", Path("dev_v1/original/few_shot")),
    ("dev_v1_expand", Path("dev_v1/expand")),
    ("dev_v1_cleaned", Path("dev_v1/cleaned")),
    ("dev_v2", Path("dev_v2")),
)

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict[str, Any]) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def validate_all_variants(results_root: Path) -> None:
    missing = [
        str(results_root / variant_path / baseline_dir / "metrics_summary.json")
        for _, variant_path in DATASET_VARIANTS
        for baseline_dir in BASELINE_DIRS
        if not (results_root / variant_path / baseline_dir / "metrics_summary.json").exists()
    ]
    if missing:
        raise FileNotFoundError(f"Missing metrics file(s): {missing}")


def build_comparison(dataset_dir: Path) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}

    for baseline_dir in BASELINE_DIRS:
        summary_path = dataset_dir / baseline_dir / "metrics_summary.json"
        if not summary_path.exists():
            raise FileNotFoundError(f"Missing metrics file: {summary_path}")
        summaries[baseline_dir] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for baseline_dir in BASELINE_DIRS:
            lang_metrics = {
                lang: summaries[baseline_dir].get((lang, mode)) for lang in LANG_ORDER
            }
            if all(metrics is None for metrics in lang_metrics.values()):
                continue

            row: dict[str, object] = {
                "mode": mode,
                "model": BASELINE_LABELS[baseline_dir],
            }

            for column, _, _ in METRICS:
                labeled_values: dict[str, float | None] = {}
                for lang in LANG_ORDER:
                    metrics = lang_metrics[lang]
                    value = metrics.get(column) if metrics else None
                    row[f"{lang}_{column}"] = value
                    labeled_values[lang] = value

                row[f"best_{column}"] = best_label(labeled_values)

            rows.append(row)

    columns = ["mode", "model"]
    for column, _, _ in METRICS:
        columns.extend([f"{lang}_{column}" for lang in LANG_ORDER])
        columns.append(f"best_{column}")

    return pd.DataFrame(rows, columns=columns)


def write_sheet(ws, df: pd.DataFrame) -> None:
    fixed_headers = ("mode", "model")
    value_subheaders = LANG_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["model"] == BASELINE_LABELS[BASELINE_DIRS[-1]]

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["model"] == BASELINE_LABELS[BASELINE_DIRS[0]] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        model_cell = ws.cell(row=row_offset, column=2, value=record["model"])
        apply_cell_style(model_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{lang}_{column}"] for lang in LANG_ORDER]
            values.append(record[f"best_{column}"])

            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill_font = (
                    label_fill(value)
                    if offset == cols_per_metric - 1 and isinstance(value, str)
                    else None
                )
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    for mode_idx in range(len(MODE_ORDER)):
        start_row = 3 + mode_idx * len(BASELINE_DIRS)
        end_row = start_row + len(BASELINE_DIRS) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/{original,expand,cleaned} and dev_v2 result folders",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/02_term_expansion_by_language_pair/report/language_comparison.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    validate_all_variants(results_root)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    row_count = 0
    for sheet_title, variant_path in DATASET_VARIANTS:
        df = build_comparison(results_root / variant_path)
        row_count = len(df)
        ws = wb.create_sheet(title=sheet_title)
        write_sheet(ws, df)

    wb.save(output_path)

    variant_names = ", ".join(name for name, _ in DATASET_VARIANTS)
    print(
        f"Wrote {len(DATASET_VARIANTS)} sheets ({variant_names}), "
        f"{row_count} rows each ({len(MODE_ORDER)} modes x {len(BASELINE_DIRS)} models), "
        f"to {output_path}"
    )


if __name__ == "__main__":
    main()
