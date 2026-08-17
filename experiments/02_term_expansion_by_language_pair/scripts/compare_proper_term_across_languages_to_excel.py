"""Compare ende, enru, and enes for proper_term mode across dev_v1 term-list variants.

Writes one styled .xlsx file (default:
``experiments/02_term_expansion_by_language_pair/report/dev_v1_proper_term_across_languages.xlsx``)
with 9 data rows: 3 term-list variants (original, expand, cleaned), each with
3 model rows (GPT, Qwen 3B, Qwen 7B). Only ``proper_term`` mode is included.
Reads ``metrics_summary.json`` from ``<variant_dir>/{gpt,qwen_3b,qwen_7b}/``.

Each value cell is colored by ranking that (model, language) combination's
value **across the 3 variants** (not across languages): red = worst variant,
yellow = middle, green = best variant. This is a different axis than the
``best`` column, which colors by the best *language* within a single row.
Matches the sibling ``dev_v1_proper_term_across_models.xlsx`` workbook
(``01_term_expansion_by_model/scripts/``).

Note: ``results/dev_v1/original/`` has no ``gpt``/``qwen_3b``/``qwen_7b``
subfolders directly — it's nested under ``no-few-shots/`` or
``with-few-shots/`` (see ``report/README.md`` §3.4.1). This script defaults
to ``with-few-shots``, which is the source of the sibling
``dev_v1_proper_term_across_models.xlsx`` workbook's numbers; override with
``--original`` if you want the ``no-few-shots`` variant instead (note that
tree only has ``proper_term`` mode evaluated, so this still works, but see
--original's help for the path).

Usage::

    python experiments/02_term_expansion_by_language_pair/scripts/compare_proper_term_across_languages_to_excel.py
    python experiments/02_term_expansion_by_language_pair/scripts/compare_proper_term_across_languages_to_excel.py --original results/dev_v1/original/no-few-shots
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LANG_ORDER = ("ende", "enru", "enes")
VARIANT_ORDER = ("original", "expand", "cleaned")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
BASELINE_LABELS = {
    "gpt": "GPT",
    "qwen_3b": "Qwen 3B",
    "qwen_7b": "Qwen 7B",
}
MODE = "proper_term"

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)

HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
RANK_FILLS = {
    "low": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "mid": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "high": PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
}
BEST_FILLS = {
    "ende": PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid"),
    "enru": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "enes": PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid"),
    "tie": PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_proper_term_metrics(summary: dict[str, Any]) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(MODE)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {column: extract_metric(metrics, spec) for column, spec, _ in METRICS}
    return by_lang


def best_language(values: dict[str, float | None]) -> str | None:
    present = {label: value for label, value in values.items() if value is not None}
    if not present:
        return None

    max_val = max(present.values())
    winners = [label for label, value in present.items() if abs(value - max_val) < 1e-9]
    if len(winners) > 1:
        return "tie"
    return winners[0]


def build_comparison(variant_dirs: dict[str, Path]) -> dict[tuple[str, str], dict[str, object]]:
    summaries: dict[tuple[str, str], dict[str, dict[str, float | None]]] = {}

    for variant, variant_dir in variant_dirs.items():
        for baseline_dir in BASELINE_DIRS:
            summary_path = variant_dir / baseline_dir / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(f"Missing metrics file: {summary_path}")
            summaries[(variant, baseline_dir)] = extract_proper_term_metrics(load_summary(summary_path))

    rows: dict[tuple[str, str], dict[str, object]] = {}
    for variant in VARIANT_ORDER:
        for baseline_dir in BASELINE_DIRS:
            lang_metrics = summaries[(variant, baseline_dir)]
            if not lang_metrics:
                continue

            row: dict[str, object] = {"data": variant, "model": BASELINE_LABELS[baseline_dir]}

            for column, _, _ in METRICS:
                labeled_values: dict[str, float | None] = {}
                for lang in LANG_ORDER:
                    value = lang_metrics.get(lang, {}).get(column)
                    row[f"{lang}_{column}"] = value
                    labeled_values[lang] = value

                row[f"best_{column}"] = best_language(labeled_values)

            rows[(variant, baseline_dir)] = row

    return rows


def variant_rank_fill(
    rows: dict[tuple[str, str], dict[str, object]], baseline_dir: str, key: str, variant: str
) -> PatternFill | None:
    """Rank a (model, language) value against its counterpart in the other 2 variants."""
    values = {
        v: rows[(v, baseline_dir)][key]
        for v in VARIANT_ORDER
        if (v, baseline_dir) in rows and rows[(v, baseline_dir)][key] is not None
    }
    if variant not in values:
        return None

    max_val = max(values.values())
    min_val = min(values.values())
    value = values[variant]
    if value == max_val:
        return RANK_FILLS["high"]
    if value == min_val:
        return RANK_FILLS["low"]
    return RANK_FILLS["mid"]


def apply_cell_style(
    cell: Cell,
    *,
    bold: bool = False,
    fill: PatternFill | None = None,
) -> None:
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER


def write_styled_excel(rows: dict[tuple[str, str], dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "proper_term"

    fixed_headers = ("data", "model")
    value_subheaders = LANG_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    row_offset = 3
    for variant in VARIANT_ORDER:
        for baseline_idx, baseline_dir in enumerate(BASELINE_DIRS):
            record = rows.get((variant, baseline_dir))
            if record is None:
                continue

            data_cell = ws.cell(row=row_offset, column=1, value=variant if baseline_idx == 0 else None)
            apply_cell_style(data_cell)

            model_cell = ws.cell(row=row_offset, column=2, value=BASELINE_LABELS[baseline_dir])
            apply_cell_style(model_cell)

            for metric_idx, (column, _, _) in enumerate(METRICS):
                start_col = metric_start_col + metric_idx * cols_per_metric
                lang_keys = [f"{lang}_{column}" for lang in LANG_ORDER]

                for offset, key in enumerate(lang_keys):
                    cell = ws.cell(row=row_offset, column=start_col + offset, value=record[key])
                    apply_cell_style(cell, fill=variant_rank_fill(rows, baseline_dir, key, variant))

                best_cell = ws.cell(
                    row=row_offset,
                    column=start_col + cols_per_metric - 1,
                    value=record[f"best_{column}"],
                )
                apply_cell_style(best_cell, fill=BEST_FILLS.get(record[f"best_{column}"]))

            row_offset += 1

    for variant_idx in range(len(VARIANT_ORDER)):
        start_row = 3 + variant_idx * len(BASELINE_DIRS)
        end_row = start_row + len(BASELINE_DIRS) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, metric_start_col + len(METRICS) * cols_per_metric):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1 result folders",
    )
    parser.add_argument(
        "--original",
        type=Path,
        default=None,
        help=(
            "Path to the dev_v1/original results (default: "
            "<results-root>/dev_v1/original/with-few-shots — "
            "override with <results-root>/dev_v1/original/no-few-shots if preferred)"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "experiments/02_term_expansion_by_language_pair/report/dev_v1_proper_term_across_languages.xlsx"
        ),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    variant_dirs = {
        "original": (args.original.resolve() if args.original else results_root / "dev_v1" / "original" / "with-few-shots"),
        "expand": results_root / "dev_v1" / "expand",
        "cleaned": results_root / "dev_v1" / "cleaned",
    }

    rows = build_comparison(variant_dirs)
    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_styled_excel(rows, output_path)

    print(
        f"Wrote {len(rows)} rows "
        f"({len(VARIANT_ORDER)} variants x {len(BASELINE_DIRS)} models) to {output_path}"
    )


if __name__ == "__main__":
    main()
