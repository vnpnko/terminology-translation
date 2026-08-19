"""Compare dev_v1/original vs dev_v2 metrics for one baseline model.

Writes one styled .xlsx file (default:
``experiments/03_dataset_comparison/report/dev_v1_original_vs_dev_v2_<baseline>_dataset_comparison.xlsx``)
with 9 data rows (mode x language). The mode column is merged per block
(no_term, proper_term, random_term). Each metric block has dev_v1_original,
dev_v2, and best columns. Reads ``metrics_summary.json`` from
``<results-root>/dev_v1/original/few_shot/<baseline>/`` (the only dev_v1/original
variant with all 3 modes) and ``<results-root>/dev_v2/<baseline>/``.

Usage::

    python experiments/03_dataset_comparison/scripts/compare_datasets_to_excel.py --baseline gpt
    python experiments/03_dataset_comparison/scripts/compare_datasets_to_excel.py --baseline qwen_3b
    python experiments/03_dataset_comparison/scripts/compare_datasets_to_excel.py --baseline qwen_7b
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    best_label,
    label_fill,
)

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
DATASET_ORDER = ("dev_v1_original", "dev_v2")
DATASET_PATHS = {
    "dev_v1_original": Path("dev_v1/original/few_shot"),
    "dev_v2": Path("dev_v2"),
}

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict[str, Any]) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def validate_all_baselines(results_root: Path) -> None:
    for dataset_label, dataset_path in DATASET_PATHS.items():
        resolved = (results_root / dataset_path).resolve()
        for baseline_dir in BASELINE_DIRS:
            summary_path = resolved / baseline_dir / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(
                    f"Missing metrics file for {dataset_label}: {summary_path}"
                )


def build_comparison(results_root: Path, baseline: str) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}

    for dataset_label, dataset_path in DATASET_PATHS.items():
        summary_path = (results_root / dataset_path / baseline / "metrics_summary.json").resolve()
        summaries[dataset_label] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for lang in LANG_ORDER:
            dataset_metrics = {
                dataset_label: summaries[dataset_label].get((lang, mode))
                for dataset_label in DATASET_ORDER
            }
            if all(metrics is None for metrics in dataset_metrics.values()):
                continue

            row: dict[str, object] = {"mode": mode, "language": lang}

            for column, _, _ in METRICS:
                labeled_values: dict[str, float | None] = {}
                for dataset_label in DATASET_ORDER:
                    metrics = dataset_metrics[dataset_label]
                    value = metrics.get(column) if metrics else None
                    row[f"{dataset_label}_{column}"] = value
                    labeled_values[dataset_label] = value

                row[f"best_{column}"] = best_label(labeled_values)

            rows.append(row)

    columns = ["mode", "language"]
    for column, _, _ in METRICS:
        columns.extend([f"{dataset_label}_{column}" for dataset_label in DATASET_ORDER])
        columns.append(f"best_{column}")

    return pd.DataFrame(rows, columns=columns)


def write_styled_excel(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "datasets"

    fixed_headers = ("mode", "language")
    value_subheaders = DATASET_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["language"] == "enes"

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["language"] == LANG_ORDER[0] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        apply_cell_style(lang_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{dataset_label}_{column}"] for dataset_label in DATASET_ORDER]
            values.append(record[f"best_{column}"])

            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill_font = (
                    label_fill(value)
                    if offset == cols_per_metric - 1 and isinstance(value, str)
                    else None
                )
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    for mode_idx in range(len(MODE_ORDER)):
        start_row = 3 + mode_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    autofit_columns(ws)

    wb.save(output_path)


def default_output_path(baseline: str, report_dir: Path) -> Path:
    return report_dir / f"dev_v1_original_vs_dev_v2_{baseline}_dataset_comparison.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--baseline",
        required=True,
        choices=BASELINE_DIRS,
        help="Which baseline model to compare datasets for",
    )
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/original and dev_v2 result folders",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Output .xlsx path (default: experiments/03_dataset_comparison/report/"
            "dev_v1_original_vs_dev_v2_<baseline>_dataset_comparison.xlsx)"
        ),
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=Path("experiments/03_dataset_comparison/report"),
        help="Report output directory when --output is not set",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    validate_all_baselines(results_root)

    output_path = (
        args.output.resolve()
        if args.output
        else default_output_path(args.baseline, args.report_dir.resolve())
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = build_comparison(results_root, args.baseline)
    write_styled_excel(df, output_path)

    print(f"Baseline: {args.baseline}")
    print(
        f"Wrote {len(df)} rows "
        f"({len(MODE_ORDER)} modes x {len(LANG_ORDER)} languages) to {output_path}"
    )


if __name__ == "__main__":
    main()
