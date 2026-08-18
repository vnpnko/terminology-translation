"""Compare dev_v1/original vs dev_v1/dictionary GPT baseline results.

Writes one styled .xlsx file (default:
``experiments/03_dataset_comparison/report/dev_v1_original_vs_dev_v1_dictionary_gpt_comparison.xlsx``).
Reads ``metrics_summary.json`` from ``<results-root>/dev_v1/original/<baseline>/``
and ``<results-root>/dev_v1/dictionary/<baseline>/``.

Usage::

    python experiments/03_dataset_comparison/scripts/compare_v1_variants_to_excel.py
    python experiments/03_dataset_comparison/scripts/compare_v1_variants_to_excel.py --baseline gpt
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
VARIANT_ORDER = ("dev_v1_original", "dev_v1_dictionary")
VARIANT_PATHS = {
    "dev_v1_original": Path("dev_v1/original"),
    "dev_v1_dictionary": Path("dev_v1/dictionary"),
}

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BEST_FILLS = {
    "dev_v1_original": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "dev_v1_dictionary": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "tie": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THICK = Side(style="medium", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict[str, Any]) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue
            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }
    return by_lang_mode


def best_variant(values: dict[str, float | None]) -> str | None:
    present = {label: value for label, value in values.items() if value is not None}
    if not present:
        return None
    max_val = max(present.values())
    winners = [label for label, value in present.items() if abs(value - max_val) < 1e-9]
    return "tie" if len(winners) > 1 else winners[0]


def build_comparison(results_root: Path, baseline: str) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}
    for variant_label, variant_path in VARIANT_PATHS.items():
        summary_path = results_root / variant_path / baseline / "metrics_summary.json"
        if not summary_path.is_file():
            raise FileNotFoundError(f"Missing: {summary_path}")
        summaries[variant_label] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []
    for mode in MODE_ORDER:
        for lang in LANG_ORDER:
            variant_metrics = {
                label: summaries[label].get((lang, mode)) for label in VARIANT_ORDER
            }
            if all(m is None for m in variant_metrics.values()):
                continue
            row: dict[str, object] = {"mode": mode, "language": lang}
            for column, _, _ in METRICS:
                labeled: dict[str, float | None] = {}
                for variant_label in VARIANT_ORDER:
                    metrics = variant_metrics[variant_label]
                    value = metrics.get(column) if metrics else None
                    row[f"{variant_label}_{column}"] = value
                    labeled[variant_label] = value
                row[f"best_{column}"] = best_variant(labeled)
            rows.append(row)

    columns = ["mode", "language"]
    for column, _, _ in METRICS:
        columns.extend([f"{v}_{column}" for v in VARIANT_ORDER])
        columns.append(f"best_{column}")
    return pd.DataFrame(rows, columns=columns)


def autofit_columns(ws, *, min_width: int = 8, max_width: int = 40, padding: int = 2) -> None:
    """Size each column from its own cell contents (openpyxl has no true AutoFit)."""
    excluded = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col > merged_range.min_col:
            excluded.add((merged_range.min_row, merged_range.min_col))

    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in excluded:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(width + padding, max_width))


def write_styled_excel(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "variants"

    fixed_headers = ("mode", "language")
    value_subheaders = VARIANT_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)
    metric_start_col = len(fixed_headers) + 1

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1
        title_cell = ws.cell(row=1, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.fill = HEADER_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            sub_cell.font = Font(bold=True)
            sub_cell.fill = HEADER_FILL
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["language"] == "enes"
        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["language"] == LANG_ORDER[0] else None,
        )
        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        for cell in (mode_cell, lang_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if thick_bottom:
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{v}_{column}"] for v in VARIANT_ORDER]
            values.append(record[f"best_{column}"])
            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if offset == cols_per_metric - 1 and isinstance(value, str):
                    cell.fill = BEST_FILLS.get(value)
                if thick_bottom:
                    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

    for mode_idx in range(len(MODE_ORDER)):
        start_row = 3 + mode_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    autofit_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--baseline", default="gpt", choices=("gpt", "qwen_3b", "qwen_7b"))
    parser.add_argument("--results-root", type=Path, default=Path("results"))
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "experiments/03_dataset_comparison/report/"
            "dev_v1_original_vs_dev_v1_dictionary_gpt_comparison.xlsx"
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    df = build_comparison(results_root, args.baseline)
    output_path = args.output.resolve()
    write_styled_excel(df, output_path)
    print(f"Wrote {len(df)} rows to {output_path}")


if __name__ == "__main__":
    main()
