"""Compare Qwen base (with few-shot) vs LoRA 1 epoch (no few-shot), for both model sizes.

Writes one .xlsx file (default:
``experiments/04_lora_finetuning/report/base_few_shots_vs_lora_1_epoch.xlsx``) with a single
sheet, one block per model size (``Qwen2.5-3B``, ``Qwen2.5-7B``, 3 rows each). Both columns
are local, driven by ``run_registry.json``:

- ``qwen_base_with_few_shots`` = ``<results-root>/<model_dir>/qwen_base/metrics_summary.json``.
- ``qwen_lora_no_few_shots`` = ``<results-root>/<model_dir>/<folder>/metrics_summary.json``,
  where ``<folder>`` is the registry's ``lora_1ep_nofs`` run.

Each value cell is colored green if it is the higher (or tied-highest) of its
qwen_base_with_few_shots/qwen_lora_no_few_shots pair, yellow otherwise — no red is used.
Same coloring convention as
``experiments/04_lora_finetuning/scripts/compare_few_shots_to_excel.py``.

Usage::

    python experiments/04_lora_finetuning/scripts/compare_base_vs_lora_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEYS = ("3B", "7B")
LEFT_LABEL = "qwen_base_with_few_shots"
RIGHT_LABEL = "qwen_lora_no_few_shots"

HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
PAIR_FILLS = {
    "low": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "high": PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def lora_1ep_nofs_run(registry: dict[str, Any], model_key: str) -> dict[str, Any]:
    runs = {run["run_id"]: run for run in registry[model_key]["runs"]}
    return runs["lora_1ep_nofs"]


def model_rows(results_root: Path, registry: dict[str, Any], model_key: str, model_dir: str) -> dict[str, tuple[list, list]]:
    base_summary = load_metrics_summary(results_root / model_dir / "qwen_base" / "metrics_summary.json")
    lora_run = lora_1ep_nofs_run(registry, model_key)
    lora_summary = load_metrics_summary(results_root / model_dir / lora_run["folder"] / "metrics_summary.json")
    return {lang: (extract_row(base_summary, lang), extract_row(lora_summary, lang)) for lang in LANG_ORDER}


def pair_fills(left_val: float | None, right_val: float | None) -> tuple[PatternFill | None, PatternFill | None]:
    """Green for the higher (or tied-highest) value of the pair, yellow for the other."""
    if left_val is None or right_val is None:
        return None, None
    max_val = max(left_val, right_val)
    left_fill = PAIR_FILLS["high"] if left_val == max_val else PAIR_FILLS["low"]
    right_fill = PAIR_FILLS["high"] if right_val == max_val else PAIR_FILLS["low"]
    return left_fill, right_fill


def apply_cell_style(cell: Cell, *, fill: PatternFill | None = None) -> None:
    cell.font = Font(bold=False)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER


def write_group_header(ws: Worksheet, start_col: int, cols_per_group: int) -> None:
    left_cell = ws.cell(row=1, column=start_col, value=LEFT_LABEL)
    apply_cell_style(left_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + cols_per_group - 1)

    right_col = start_col + cols_per_group
    right_cell = ws.cell(row=1, column=right_col, value=RIGHT_LABEL)
    apply_cell_style(right_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=right_col, end_row=1, end_column=right_col + cols_per_group - 1)

    for group_start in (start_col, right_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)


def write_data_row(ws: Worksheet, row: int, start_col: int, left_values: list, right_values: list) -> None:
    cols_per_group = len(METRICS)
    for offset in range(cols_per_group):
        left_fill, right_fill = pair_fills(left_values[offset], right_values[offset])
        left_cell = ws.cell(row=row, column=start_col + offset, value=left_values[offset])
        apply_cell_style(left_cell, fill=left_fill)
        right_cell = ws.cell(row=row, column=start_col + cols_per_group + offset, value=right_values[offset])
        apply_cell_style(right_cell, fill=right_fill)


def write_sheet(ws: Worksheet, rows_by_model: dict[str, dict[str, tuple[list, list]]]) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 3

    model_header = ws.cell(row=1, column=1, value="model")
    apply_cell_style(model_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    lang_header = ws.cell(row=1, column=2, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    write_group_header(ws, start_col=data_start_col, cols_per_group=cols_per_group)

    row = 3
    for model_dir, rows_by_lang in rows_by_model.items():
        block_start_row = row
        for lang in LANG_ORDER:
            lang_cell = ws.cell(row=row, column=2, value=lang)
            apply_cell_style(lang_cell)
            left_values, right_values = rows_by_lang[lang]
            write_data_row(ws, row, start_col=data_start_col, left_values=left_values, right_values=right_values)
            row += 1

        block_cell = ws.cell(row=block_start_row, column=1, value=model_dir)
        apply_cell_style(block_cell)
        ws.merge_cells(start_row=block_start_row, start_column=1, end_row=row - 1, end_column=1)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col_idx in range(data_start_col, data_start_col + 2 * cols_per_group):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/04_lora_finetuning/results"),
        help="Root directory containing Qwen2.5-3B/, Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/04_lora_finetuning/scripts/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/04_lora_finetuning/report/base_few_shots_vs_lora_1_epoch.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any]) -> None:
    required = []
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        required.append(results_root / model_dir / "qwen_base" / "metrics_summary.json")
        lora_run = lora_1ep_nofs_run(registry, model_key)
        required.append(results_root / model_dir / lora_run["folder"] / "metrics_summary.json")

    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry)

    rows_by_model = {}
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        rows_by_model[model_dir] = model_rows(results_root, registry, model_key, model_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    write_sheet(ws, rows_by_model)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 1 sheet ({len(MODEL_KEYS)} models x {len(LANG_ORDER)} languages) to {output_path}")


if __name__ == "__main__":
    main()
