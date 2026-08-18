"""Compare the best Qwen 7B LoRA config vs GPT-4o-mini few-shot.

Writes one .xlsx file (default: ``experiments/04_lora_finetuning/report/best_models.xlsx``)
with a single sheet, one row per language pair. Both columns are local:

- LoRA side: ``<results-root>/Qwen2.5-7B/<folder>/metrics_summary.json``, where ``<folder>``
  is ``run_registry.json``'s entry for ``--lora-run`` (default ``lora_2_epoch_zero_shot`` —
  chosen over 3 epochs: training loss keeps dropping sharply at 3 epochs while held-out
  BLEU/chrF plateau or regress, an overfitting signature; only Term Accuracy keeps improving
  at 3 epochs).
- GPT side: ``<results-root>/gpt_base/metrics_summary.json``.

Each value cell is colored green if it is the higher (or tied-highest) of its
LoRA/GPT pair, yellow otherwise — no red is used. Same coloring convention as
``experiments/04_lora_finetuning/scripts/compare_few_shots_to_excel.py``.

Usage::

    python experiments/04_lora_finetuning/scripts/compare_best_models_to_excel.py
    python experiments/04_lora_finetuning/scripts/compare_best_models_to_excel.py --lora-run lora_3_epoch_zero_shot
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEY = "7B"
RIGHT_LABEL = "GPT-4o-mini"

HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
PAIR_FILLS = {
    "low": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "high": PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def lora_run(registry: dict[str, Any], run_id: str) -> dict[str, Any]:
    runs = {run["run_id"]: run for run in registry[MODEL_KEY]["runs"]}
    return runs[run_id]


def best_rows(results_root: Path, registry: dict[str, Any], run_id: str) -> dict[str, tuple[list, list]]:
    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, run_id)
    lora_summary = load_metrics_summary(results_root / model_dir / run["folder"] / "metrics_summary.json")
    gpt_summary = load_metrics_summary(results_root / "gpt_base" / "metrics_summary.json")
    return {lang: (extract_row(lora_summary, lang), extract_row(gpt_summary, lang)) for lang in LANG_ORDER}


def pair_fills(left_val: float | None, right_val: float | None) -> tuple[PatternFill | None, PatternFill | None]:
    """Green for the higher (or tied-highest) value of the pair, yellow for the other."""
    if left_val is None or right_val is None:
        return None, None
    max_val = max(left_val, right_val)
    left_fill = PAIR_FILLS["high"] if left_val == max_val else PAIR_FILLS["low"]
    right_fill = PAIR_FILLS["high"] if right_val == max_val else PAIR_FILLS["low"]
    return left_fill, right_fill


def apply_cell_style(cell: Cell, *, fill: PatternFill | None = None) -> None:
    cell.font = Font(bold=False)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER


def autofit_columns(ws: Worksheet, *, min_width: int = 8, max_width: int = 40, padding: int = 2) -> None:
    """Size each column from its own cell contents (openpyxl has no true AutoFit)."""
    excluded = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col > merged_range.min_col:
            excluded.add((merged_range.min_row, merged_range.min_col))

    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in excluded:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(width + padding, max_width))


def write_sheet(ws: Worksheet, rows_by_lang: dict[str, tuple[list, list]], left_label: str) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 2

    lang_header = ws.cell(row=1, column=1, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    left_cell = ws.cell(row=1, column=data_start_col, value=left_label)
    apply_cell_style(left_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=data_start_col, end_row=1, end_column=data_start_col + cols_per_group - 1)

    right_col = data_start_col + cols_per_group
    right_cell = ws.cell(row=1, column=right_col, value=RIGHT_LABEL)
    apply_cell_style(right_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=right_col, end_row=1, end_column=right_col + cols_per_group - 1)

    for group_start in (data_start_col, right_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)

    for row_offset, lang in enumerate(LANG_ORDER, start=3):
        lang_cell = ws.cell(row=row_offset, column=1, value=lang)
        apply_cell_style(lang_cell)
        left_values, right_values = rows_by_lang[lang]

        for offset in range(cols_per_group):
            left_fill, right_fill = pair_fills(left_values[offset], right_values[offset])
            left_cell = ws.cell(row=row_offset, column=data_start_col + offset, value=left_values[offset])
            apply_cell_style(left_cell, fill=left_fill)
            right_cell = ws.cell(row=row_offset, column=right_col + offset, value=right_values[offset])
            apply_cell_style(right_cell, fill=right_fill)

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/04_lora_finetuning/results"),
        help="Root directory containing gpt_base/ and Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/04_lora_finetuning/scripts/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--lora-run",
        default="lora_2_epoch_zero_shot",
        help="run_registry.json run_id (7B) to use as the 'best' LoRA config",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/04_lora_finetuning/report/best_models.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any], run_id: str) -> None:
    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, run_id)
    required = [
        results_root / model_dir / run["folder"] / "metrics_summary.json",
        results_root / "gpt_base" / "metrics_summary.json",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry, args.lora_run)

    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, args.lora_run)
    num_epochs = run["num_epochs"]
    left_label = f"qwen_lora_7b_zero_shot_{num_epochs}_epochs"

    rows_by_lang = best_rows(results_root, registry, args.lora_run)

    wb = Workbook()
    ws = wb.active
    ws.title = "best_models"
    write_sheet(ws, rows_by_lang, left_label)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 1 sheet ({len(LANG_ORDER)} languages, {model_dir}/{run['folder']} vs GPT) to {output_path}")


if __name__ == "__main__":
    main()
