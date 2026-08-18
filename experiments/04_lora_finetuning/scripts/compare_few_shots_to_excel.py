"""Compare no-few-shot vs few-shot for GPT-4o-mini, Qwen base, and Qwen LoRA (1 epoch).

Writes one .xlsx file (default: ``experiments/04_lora_finetuning/report/few_shots_ablation.xlsx``)
with 3 sheets: ``GPT-4o-mini``, ``Qwen2.5-3B``, ``Qwen2.5-7B``. The GPT sheet has one row per
language pair; the Qwen sheets stack a ``qwen_base`` block and a ``qwen_lora`` block (3 rows
each), matching the two differently-scoped few-shot experiments documented in
``report/README.md`` §3.4.1:

- **Baseline-level** (GPT, Qwen base): ``few_shots`` reads the LoRA experiment's own local
  baseline run (``<results-root>/gpt_base/`` or ``<results-root>/<model_dir>/qwen_base/``);
  ``no_shots`` reads the separate, shared ``<baseline-results-root>/dev_v1/original/no-few-shots/
  {gpt,qwen_3b,qwen_7b}/`` tree (not tracked in ``run_registry.json``).
- **LoRA-level** (``qwen_lora``): both columns are local, driven by ``run_registry.json``'s
  ``lora_1ep_nofs``/``lora_1ep_fs`` runs.

Each value cell is colored green if it is the higher (or tied-highest) of its no_shots/
few_shots pair, yellow otherwise — no red is used.

Usage::

    python experiments/04_lora_finetuning/scripts/compare_few_shots_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEYS = ("3B", "7B")
BASELINE_DIRS = {"3B": "qwen_3b", "7B": "qwen_7b"}

HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
PAIR_FILLS = {
    "low": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "high": PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def lora_1ep_runs(registry: dict[str, Any], model_key: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """Return (no_few_shot_run, few_shot_run) for 1-epoch LoRA runs."""
    runs = {run["run_id"]: run for run in registry[model_key]["runs"]}
    return runs["lora_1ep_nofs"], runs["lora_1ep_fs"]


def gpt_rows(results_root: Path, baseline_results_root: Path) -> dict[str, tuple[list, list]]:
    no_shots_summary = load_metrics_summary(
        baseline_results_root / "dev_v1" / "original" / "no-few-shots" / "gpt" / "metrics_summary.json"
    )
    few_shots_summary = load_metrics_summary(results_root / "gpt_base" / "metrics_summary.json")
    return {
        lang: (extract_row(no_shots_summary, lang), extract_row(few_shots_summary, lang)) for lang in LANG_ORDER
    }


def qwen_base_rows(
    results_root: Path, baseline_results_root: Path, model_dir: str, baseline_dir: str
) -> dict[str, tuple[list, list]]:
    no_shots_summary = load_metrics_summary(
        baseline_results_root / "dev_v1" / "original" / "no-few-shots" / baseline_dir / "metrics_summary.json"
    )
    few_shots_summary = load_metrics_summary(results_root / model_dir / "qwen_base" / "metrics_summary.json")
    return {
        lang: (extract_row(no_shots_summary, lang), extract_row(few_shots_summary, lang)) for lang in LANG_ORDER
    }


def qwen_lora_rows(results_root: Path, registry: dict[str, Any], model_key: str, model_dir: str) -> dict[str, tuple[list, list]]:
    no_shots_run, few_shots_run = lora_1ep_runs(registry, model_key)
    no_shots_summary = load_metrics_summary(results_root / model_dir / no_shots_run["folder"] / "metrics_summary.json")
    few_shots_summary = load_metrics_summary(results_root / model_dir / few_shots_run["folder"] / "metrics_summary.json")
    return {
        lang: (extract_row(no_shots_summary, lang), extract_row(few_shots_summary, lang)) for lang in LANG_ORDER
    }


def pair_fills(no_shots_val: float | None, few_shots_val: float | None) -> tuple[PatternFill | None, PatternFill | None]:
    """Green for the higher (or tied-highest) value of the pair, yellow for the other."""
    if no_shots_val is None or few_shots_val is None:
        return None, None
    max_val = max(no_shots_val, few_shots_val)
    no_fill = PAIR_FILLS["high"] if no_shots_val == max_val else PAIR_FILLS["low"]
    few_fill = PAIR_FILLS["high"] if few_shots_val == max_val else PAIR_FILLS["low"]
    return no_fill, few_fill


def apply_cell_style(cell: Cell, *, fill: PatternFill | None = None) -> None:
    cell.font = Font(bold=False)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER


def write_group_header(ws: Worksheet, start_col: int, cols_per_group: int) -> None:
    no_shots_cell = ws.cell(row=1, column=start_col, value="no_shots")
    apply_cell_style(no_shots_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + cols_per_group - 1)

    few_shots_col = start_col + cols_per_group
    few_shots_cell = ws.cell(row=1, column=few_shots_col, value="few_shots")
    apply_cell_style(few_shots_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=few_shots_col, end_row=1, end_column=few_shots_col + cols_per_group - 1)

    for group_start in (start_col, few_shots_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)


def write_data_row(ws: Worksheet, row: int, start_col: int, no_shots_values: list, few_shots_values: list) -> None:
    cols_per_group = len(METRICS)
    for offset in range(cols_per_group):
        no_fill, few_fill = pair_fills(no_shots_values[offset], few_shots_values[offset])
        no_cell = ws.cell(row=row, column=start_col + offset, value=no_shots_values[offset])
        apply_cell_style(no_cell, fill=no_fill)
        few_cell = ws.cell(row=row, column=start_col + cols_per_group + offset, value=few_shots_values[offset])
        apply_cell_style(few_cell, fill=few_fill)


def write_gpt_sheet(ws: Worksheet, rows_by_lang: dict[str, tuple[list, list]]) -> None:
    cols_per_group = len(METRICS)

    lang_header = ws.cell(row=1, column=1, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    write_group_header(ws, start_col=2, cols_per_group=cols_per_group)

    for row_offset, lang in enumerate(LANG_ORDER, start=3):
        lang_cell = ws.cell(row=row_offset, column=1, value=lang)
        apply_cell_style(lang_cell)
        no_shots_values, few_shots_values = rows_by_lang[lang]
        write_data_row(ws, row_offset, start_col=2, no_shots_values=no_shots_values, few_shots_values=few_shots_values)

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, 2 + 2 * cols_per_group):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def write_qwen_sheet(
    ws: Worksheet,
    model_dir: str,
    base_rows_by_lang: dict[str, tuple[list, list]],
    lora_rows_by_lang: dict[str, tuple[list, list]],
) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 3

    model_header = ws.cell(row=1, column=1, value=model_dir)
    apply_cell_style(model_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    lang_header = ws.cell(row=1, column=2, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    write_group_header(ws, start_col=data_start_col, cols_per_group=cols_per_group)

    row = 3
    for block_label, rows_by_lang in (("qwen_base", base_rows_by_lang), ("qwen_lora", lora_rows_by_lang)):
        block_start_row = row
        for lang in LANG_ORDER:
            lang_cell = ws.cell(row=row, column=2, value=lang)
            apply_cell_style(lang_cell)
            no_shots_values, few_shots_values = rows_by_lang[lang]
            write_data_row(ws, row, start_col=data_start_col, no_shots_values=no_shots_values, few_shots_values=few_shots_values)
            row += 1

        block_cell = ws.cell(row=block_start_row, column=1, value=block_label)
        apply_cell_style(block_cell)
        ws.merge_cells(start_row=block_start_row, start_column=1, end_row=row - 1, end_column=1)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col_idx in range(data_start_col, data_start_col + 2 * cols_per_group):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/04_lora_finetuning/results"),
        help="Root directory containing gpt_base/ and Qwen2.5-3B/, Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--baseline-results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/original/no-few-shots/{gpt,qwen_3b,qwen_7b}/",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/04_lora_finetuning/scripts/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/04_lora_finetuning/report/few_shots_ablation.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any]) -> None:
    required = [results_root / "gpt_base" / "metrics_summary.json"]
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        required.append(results_root / model_dir / "qwen_base" / "metrics_summary.json")
        no_shots_run, few_shots_run = lora_1ep_runs(registry, model_key)
        required.append(results_root / model_dir / no_shots_run["folder"] / "metrics_summary.json")
        required.append(results_root / model_dir / few_shots_run["folder"] / "metrics_summary.json")

    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    baseline_results_root = args.baseline_results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry)

    wb = Workbook()
    wb.remove(wb.active)

    gpt_ws = wb.create_sheet(title="GPT-4o-mini")
    write_gpt_sheet(gpt_ws, gpt_rows(results_root, baseline_results_root))

    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        baseline_dir = BASELINE_DIRS[model_key]
        base_rows = qwen_base_rows(results_root, baseline_results_root, model_dir, baseline_dir)
        lora_rows = qwen_lora_rows(results_root, registry, model_key, model_dir)
        ws = wb.create_sheet(title=model_dir)
        write_qwen_sheet(ws, model_dir, base_rows, lora_rows)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 3 sheets (GPT-4o-mini, Qwen2.5-3B, Qwen2.5-7B) to {output_path}")


if __name__ == "__main__":
    main()
