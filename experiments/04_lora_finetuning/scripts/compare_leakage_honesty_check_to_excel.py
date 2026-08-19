"""Build the data-leakage honesty-check workbook (no_overlap_data vs overlap_data test subsets).

Writes one .xlsx file (default:
``experiments/04_lora_finetuning/report/leakage_honesty_check.xlsx``) with a single sheet
(``overlap_vs_no_overlap_data``), an ``overlap_data`` and ``no_overlap_data`` column group
(5 metrics each — named for the actual split criterion, ≥50% token containment with the
training set, rather than a "bad"/"good" value judgment), and 3 stacked 3-row
(``ende``/``enes``/``enru``) model blocks, in this order:

1. ``qwen_base`` (untrained) — control.
2. ``gpt`` (closed model, never exposed to dev_v2 training data) — control.
3. ``qwen_lora`` (trained) — the model under test.

Controls come first so the reader sees the overlap-vs-no-overlap gap grow with training
exposure: qwen_base's gap is tiny (~0.8 BLEU), GPT's is a bit larger (~3.2, inherent
sentence difficulty), and qwen_lora's is much larger (~9.3) — evidence of leakage inflation
on top of the difficulty confound, not instead of it.

Reads ``metrics_summary.json`` from
``<results-root>/{gpt_base,<model_dir>/qwen_base,<model_dir>/<lora folder>}/test_cleaned_by_sentences/{overlap,no_overlap}/``.

Each value cell is colored green if it is the higher of its overlap_data/no_overlap_data
pair, red if it is the lower, and yellow if the pair ties (shared Good/Bad/Neutral
convention, see ``src/analysis/excel_style.py``). Same coloring convention as
``experiments/04_lora_finetuning/scripts/compare_few_shots_to_excel.py``.

Usage::

    python experiments/04_lora_finetuning/scripts/compare_leakage_honesty_check_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import apply_cell_style, autofit_columns, rank_fills  # noqa: E402

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEY = "7B"

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def lora_run(registry: dict[str, Any], run_id: str) -> dict[str, Any]:
    runs = {run["run_id"]: run for run in registry[MODEL_KEY]["runs"]}
    return runs[run_id]


def subset_rows(base_dir: Path, subset: str) -> dict[str, list[float | None]]:
    summary = load_metrics_summary(base_dir / "test_cleaned_by_sentences" / subset / "metrics_summary.json")
    return {lang: extract_row(summary, lang) for lang in LANG_ORDER}


def write_group_header(ws: Worksheet, start_col: int, cols_per_group: int) -> None:
    bad_cell = ws.cell(row=1, column=start_col, value="overlap_data")
    apply_cell_style(bad_cell)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + cols_per_group - 1)

    good_col = start_col + cols_per_group
    good_cell = ws.cell(row=1, column=good_col, value="no_overlap_data")
    apply_cell_style(good_cell)
    ws.merge_cells(start_row=1, start_column=good_col, end_row=1, end_column=good_col + cols_per_group - 1)

    for group_start in (start_col, good_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell)


def write_block(ws: Worksheet, start_row: int, label: str, bad: dict[str, list], good: dict[str, list]) -> int:
    lang_col = 2
    data_start_col = 3
    cols_per_group = len(METRICS)

    row = start_row
    for lang in LANG_ORDER:
        lang_cell = ws.cell(row=row, column=lang_col, value=lang)
        apply_cell_style(lang_cell)
        for offset in range(cols_per_group):
            fills = rank_fills({"bad": bad[lang][offset], "good": good[lang][offset]})
            bad_fill_font = fills.get("bad")
            good_fill_font = fills.get("good")
            bad_cell = ws.cell(row=row, column=data_start_col + offset, value=bad[lang][offset])
            apply_cell_style(
                bad_cell,
                fill=bad_fill_font[0] if bad_fill_font else None,
                font=bad_fill_font[1] if bad_fill_font else None,
            )
            good_cell = ws.cell(row=row, column=data_start_col + cols_per_group + offset, value=good[lang][offset])
            apply_cell_style(
                good_cell,
                fill=good_fill_font[0] if good_fill_font else None,
                font=good_fill_font[1] if good_fill_font else None,
            )
        row += 1

    block_cell = ws.cell(row=start_row, column=1, value=label)
    apply_cell_style(block_cell)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)

    return row


def write_sheet(ws: Worksheet, blocks: list[tuple[str, dict, dict]]) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 3

    model_cell = ws.cell(row=2, column=1, value="model")
    apply_cell_style(model_cell)

    lang_cell = ws.cell(row=2, column=2, value="lang_pair")
    apply_cell_style(lang_cell)

    write_group_header(ws, data_start_col, cols_per_group)

    row = 3
    for label, bad, good in blocks:
        row = write_block(ws, row, label, bad, good)

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/04_lora_finetuning/results"),
        help="Root directory containing gpt_base/ and Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/04_lora_finetuning/scripts/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--lora-run",
        default="lora_2_epoch_zero_shot",
        help="run_registry.json run_id (7B) to use for the LoRA side",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/04_lora_finetuning/report/leakage_honesty_check.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, model_dir: str, lora_folder: str) -> None:
    required = []
    for base_dir in (results_root / "gpt_base", results_root / model_dir / "qwen_base", results_root / model_dir / lora_folder):
        for subset in ("overlap", "no_overlap"):
            required.append(base_dir / "test_cleaned_by_sentences" / subset / "metrics_summary.json")

    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, args.lora_run)

    validate_required(results_root, model_dir, run["folder"])

    gpt_dir = results_root / "gpt_base"
    base_dir = results_root / model_dir / "qwen_base"
    lora_dir = results_root / model_dir / run["folder"]

    gpt_bad, gpt_good = subset_rows(gpt_dir, "overlap"), subset_rows(gpt_dir, "no_overlap")
    base_bad, base_good = subset_rows(base_dir, "overlap"), subset_rows(base_dir, "no_overlap")
    lora_bad, lora_good = subset_rows(lora_dir, "overlap"), subset_rows(lora_dir, "no_overlap")

    wb = Workbook()
    ws = wb.active
    ws.title = "overlap_vs_no_overlap_data"
    write_sheet(
        ws,
        [
            ("qwen_base", base_bad, base_good),
            ("gpt", gpt_bad, gpt_good),
            ("qwen_lora", lora_bad, lora_good),
        ],
    )

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 1 sheet (3 models x {len(LANG_ORDER)} languages) to {output_path}")


if __name__ == "__main__":
    main()
