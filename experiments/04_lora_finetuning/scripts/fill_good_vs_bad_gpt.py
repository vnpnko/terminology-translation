"""Fill good-vs-bad comparison blocks in comparisons.xlsx from gpt_base metrics."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "report" / "comparisons.xlsx"
SHEET = "good vs bad data"
QWEN_BASE = ROOT / "results" / "Qwen2.5-7B" / "qwen_base"
BLOCKS = {
    "gpt_bad_data vs gpt_good_data": {
        "lang_rows": {"ende": 3, "enes": 4, "enru": 5},
        "bad_path": ROOT / "results" / "gpt_base" / "test_cleaned_by_sentences" / "data_bad" / "metrics_summary.json",
        "good_path": ROOT / "results" / "gpt_base" / "test_cleaned_by_sentences" / "data_good" / "metrics_summary.json",
    },
    "qwen_base_bad_data vs qwen_base_good_data": {
        "lang_rows": {"ende": 10, "enes": 11, "enru": 12},
        "bad_path": QWEN_BASE / "test_cleaned_by_sentences" / "data_bad" / "metrics_summary.json",
        "good_path": QWEN_BASE / "metrics_summary.json",
    },
}
GREEN_FILL = PatternFill(fill_type="solid", fgColor="00B050")
RED_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
METRIC_COLS = range(2, 7)


def extract_row(data: dict, lang_key: str) -> list[float]:
    metrics = data["languages"][lang_key]["modes"]["proper_term"]["metrics"]
    consistency = metrics["terminology_consistency"]
    return [
        round(metrics["bleu"], 2),
        round(metrics["chrf"], 2),
        round(metrics["terminology_accuracy"]["avg_ratio_pct"], 2),
        round(consistency["macro_avg_consistency"], 4),
        round(consistency["weighted_avg_consistency"], 4),
    ]


def color_block(ws, lang_rows: dict[str, int]) -> None:
    for row in lang_rows.values():
        for bad_col in METRIC_COLS:
            good_col = bad_col + 5
            bad_val = ws.cell(row=row, column=bad_col).value
            good_val = ws.cell(row=row, column=good_col).value
            if bad_val is None or good_val is None:
                continue
            if bad_val > good_val:
                ws.cell(row=row, column=bad_col).fill = GREEN_FILL
                ws.cell(row=row, column=good_col).fill = RED_FILL
            elif bad_val < good_val:
                ws.cell(row=row, column=bad_col).fill = RED_FILL
                ws.cell(row=row, column=good_col).fill = GREEN_FILL
            else:
                ws.cell(row=row, column=bad_col).fill = YELLOW_FILL
                ws.cell(row=row, column=good_col).fill = YELLOW_FILL


def fill_block(ws, bad: dict, good: dict, lang_rows: dict[str, int]) -> None:
    for lang_key, row in lang_rows.items():
        for col_offset, value in enumerate(extract_row(bad, lang_key), start=2):
            ws.cell(row=row, column=col_offset, value=value)
        for col_offset, value in enumerate(extract_row(good, lang_key), start=7):
            ws.cell(row=row, column=col_offset, value=value)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    for block_name, block in BLOCKS.items():
        bad = json.loads(block["bad_path"].read_text(encoding="utf-8"))
        good = json.loads(block["good_path"].read_text(encoding="utf-8"))
        lang_rows = block["lang_rows"]
        fill_block(ws, bad, good, lang_rows)
        color_block(ws, lang_rows)
        rows = ", ".join(str(r) for r in lang_rows.values())
        print(f"Updated {block_name} ({SHEET} rows {rows}).")

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    main()
