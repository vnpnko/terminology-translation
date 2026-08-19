"""Compare dev_v1/original vs dev_v2 metrics for all three baseline models.

Writes one styled .xlsx file (default:
``experiments/dataset_comparison/report/dataset_comparison.xlsx``)
with one sheet per baseline (``gpt``, ``qwen_3b``, ``qwen_7b``), each with 9
data rows (mode x language). The mode column is merged per block (no_term,
proper_term, random_term). Each metric block has dev_v1 and dev_v2 columns,
colored with the shared Good/Bad/Neutral convention (see
``src/analysis/excel_style.py``): green for the higher value, red for the
lower, yellow if they're within 1% of each other. Reads ``metrics_summary.json``
from ``<results-root>/dev_v1/original/few_shot/<baseline>/`` (the only dev_v1/original
variant with all 3 modes) and ``<results-root>/dev_v2/original/<baseline>/``.

Usage::

    python experiments/dataset_comparison/scripts/compare_datasets_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
DATASET_ORDER = ("dev_v1", "dev_v2")
DATASET_PATHS = {
    "dev_v1": Path("dev_v1/original/few_shot"),
    "dev_v2": Path("dev_v2/original"),
}

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict[str, Any]) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def validate_all_baselines(results_root: Path) -> None:
    for dataset_label, dataset_path in DATASET_PATHS.items():
        resolved = (results_root / dataset_path).resolve()
        for baseline_dir in BASELINE_DIRS:
            summary_path = resolved / baseline_dir / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(
                    f"Missing metrics file for {dataset_label}: {summary_path}"
                )


def build_comparison(results_root: Path, baseline: str) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}

    for dataset_label, dataset_path in DATASET_PATHS.items():
        summary_path = (results_root / dataset_path / baseline / "metrics_summary.json").resolve()
        summaries[dataset_label] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for lang in LANG_ORDER:
            dataset_metrics = {
                dataset_label: summaries[dataset_label].get((lang, mode))
                for dataset_label in DATASET_ORDER
            }
            if all(metrics is None for metrics in dataset_metrics.values()):
                continue

            row: dict[str, object] = {"mode": mode, "language": lang}

            for column, _, _ in METRICS:
                for dataset_label in DATASET_ORDER:
                    metrics = dataset_metrics[dataset_label]
                    value = metrics.get(column) if metrics else None
                    row[f"{dataset_label}_{column}"] = value

            rows.append(row)

    columns = ["mode", "language"]
    for column, _, _ in METRICS:
        columns.extend([f"{dataset_label}_{column}" for dataset_label in DATASET_ORDER])

    return pd.DataFrame(rows, columns=columns)


def write_sheet(ws, df: pd.DataFrame) -> None:
    fixed_headers = ("mode", "language")
    value_subheaders = DATASET_ORDER
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["language"] == "enes"

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["language"] == LANG_ORDER[0] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        apply_cell_style(lang_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{dataset_label}_{column}"] for dataset_label in DATASET_ORDER]
            fills = rank_fills(dict(zip(DATASET_ORDER, values)))

            for offset, (dataset_label, value) in enumerate(zip(DATASET_ORDER, values)):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill_font = fills.get(dataset_label)
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    for mode_idx in range(len(MODE_ORDER)):
        start_row = 3 + mode_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/original and dev_v2 result folders",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "experiments/dataset_comparison/report/"
            "dataset_comparison.xlsx"
        ),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    validate_all_baselines(results_root)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    row_count = 0
    for baseline_dir in BASELINE_DIRS:
        df = build_comparison(results_root, baseline_dir)
        row_count = len(df)
        ws = wb.create_sheet(title=baseline_dir)
        write_sheet(ws, df)

    wb.save(output_path)

    print(
        f"Wrote {len(BASELINE_DIRS)} sheets ({', '.join(BASELINE_DIRS)}), "
        f"{row_count} rows each ({len(MODE_ORDER)} modes x {len(LANG_ORDER)} languages), "
        f"to {output_path}"
    )


if __name__ == "__main__":
    main()
