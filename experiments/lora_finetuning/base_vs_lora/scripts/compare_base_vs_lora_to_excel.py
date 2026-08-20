"""Compare Qwen base (few-shot) vs LoRA 1 epoch (zero-shot), for both model sizes.

Writes one .xlsx file (default:
``experiments/lora_finetuning/base_vs_lora/report/base_few_shot_vs_lora_zero_shot_1_epoch.xlsx``) with a
single sheet, one block per model size (``Qwen2.5-3B``, ``Qwen2.5-7B``, 3 rows each). Both
columns are local, driven by ``run_registry.json``:

- ``qwen_base_few_shot`` = ``<results-root>/<model_dir>/qwen_base/metrics_summary.json``.
- ``qwen_lora_zero_shot`` = ``<results-root>/<model_dir>/<folder>/metrics_summary.json``,
  where ``<folder>`` is the registry's ``lora_1_epoch_zero_shot`` run.

Each value cell is colored by the shared Good/Bad/Neutral convention (see
``shared/lib/analysis/excel_style.py``): green for the higher of its
qwen_base_few_shot/qwen_lora_zero_shot pair, red for the lower, yellow if equal.

Usage::

    python experiments/lora_finetuning/base_vs_lora/scripts/compare_base_vs_lora_to_excel.py
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "experiments" / "lora_finetuning" / "shared" / "scripts"))

from compare_common import (  # noqa: E402
    METRICS,
    RUN_LORA_1_EPOCH_ZERO_SHOT,
    extract_row,
    load_metrics_summary,
    load_registry,
    write_group_header,
)
from shared.lib.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEYS = ("3B", "7B")
LEFT_LABEL = "qwen_base_few_shot"
RIGHT_LABEL = "qwen_lora_zero_shot"


def lora_1_epoch_zero_shot_run(registry: dict[str, Any], model_key: str) -> dict[str, Any]:
    runs = {run["run_id"]: run for run in registry[model_key]["runs"]}
    return runs[RUN_LORA_1_EPOCH_ZERO_SHOT]


def model_rows(results_root: Path, registry: dict[str, Any], model_key: str, model_dir: str) -> dict[str, tuple[list, list]]:
    base_summary = load_metrics_summary(results_root / model_dir / "qwen_base" / "metrics_summary.json")
    lora_run = lora_1_epoch_zero_shot_run(registry, model_key)
    lora_summary = load_metrics_summary(results_root / model_dir / lora_run["folder"] / "metrics_summary.json")
    return {lang: (extract_row(base_summary, lang), extract_row(lora_summary, lang)) for lang in LANG_ORDER}


def write_data_row(ws: Worksheet, row: int, start_col: int, left_values: list, right_values: list) -> None:
    cols_per_group = len(METRICS)
    for offset in range(cols_per_group):
        fills = rank_fills({"left": left_values[offset], "right": right_values[offset]})
        left_fill_font = fills.get("left")
        right_fill_font = fills.get("right")

        left_cell = ws.cell(row=row, column=start_col + offset, value=left_values[offset])
        apply_cell_style(
            left_cell,
            fill=left_fill_font[0] if left_fill_font else None,
            font=left_fill_font[1] if left_fill_font else None,
        )
        right_cell = ws.cell(row=row, column=start_col + cols_per_group + offset, value=right_values[offset])
        apply_cell_style(
            right_cell,
            fill=right_fill_font[0] if right_fill_font else None,
            font=right_fill_font[1] if right_fill_font else None,
        )


def write_sheet(ws: Worksheet, rows_by_model: dict[str, dict[str, tuple[list, list]]]) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 3

    model_header = ws.cell(row=1, column=1, value="model")
    apply_cell_style(model_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    lang_header = ws.cell(row=1, column=2, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    write_group_header(ws, data_start_col, cols_per_group, LEFT_LABEL, RIGHT_LABEL, fill=HEADER_FILL)

    row = 3
    for model_dir, rows_by_lang in rows_by_model.items():
        block_start_row = row
        for lang in LANG_ORDER:
            lang_cell = ws.cell(row=row, column=2, value=lang)
            apply_cell_style(lang_cell)
            left_values, right_values = rows_by_lang[lang]
            write_data_row(ws, row, start_col=data_start_col, left_values=left_values, right_values=right_values)
            row += 1

        block_cell = ws.cell(row=block_start_row, column=1, value=model_dir)
        apply_cell_style(block_cell)
        ws.merge_cells(start_row=block_start_row, start_column=1, end_row=row - 1, end_column=1)

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/lora_finetuning/shared/results"),
        help="Root directory containing Qwen2.5-3B/, Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/lora_finetuning/shared/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/lora_finetuning/base_vs_lora/report/base_few_shot_vs_lora_zero_shot_1_epoch.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any]) -> None:
    required = []
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        required.append(results_root / model_dir / "qwen_base" / "metrics_summary.json")
        lora_run = lora_1_epoch_zero_shot_run(registry, model_key)
        required.append(results_root / model_dir / lora_run["folder"] / "metrics_summary.json")

    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry)

    rows_by_model = {}
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        rows_by_model[model_dir] = model_rows(results_root, registry, model_key, model_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "base_vs_lora"
    write_sheet(ws, rows_by_model)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 1 sheet ({len(MODEL_KEYS)} models x {len(LANG_ORDER)} languages) to {output_path}")


if __name__ == "__main__":
    main()
