"""Compare the best Qwen 7B LoRA config vs GPT-4o-mini few-shot.

Writes one .xlsx file (default: ``experiments/lora_finetuning/best_models/report/best_models.xlsx``)
with a single sheet, one row per language pair. Both columns are local:

- LoRA side: ``<results-root>/Qwen2.5-7B/<folder>/metrics_summary.json``, where ``<folder>``
  is ``run_registry.json``'s entry for ``--lora-run`` (default ``lora_2_epoch_zero_shot`` —
  chosen over 3 epochs: training loss keeps dropping sharply at 3 epochs while held-out
  BLEU/chrF plateau or regress, an overfitting signature; only Term Accuracy keeps improving
  at 3 epochs).
- GPT side: ``<results-root>/gpt_base/metrics_summary.json``.

Each value cell is colored by the shared Good/Bad/Neutral convention (see
``src/analysis/excel_style.py``): green for the higher of its LoRA/GPT pair,
red for the lower, yellow if equal.

Usage::

    python experiments/lora_finetuning/best_models/scripts/compare_best_models_to_excel.py
    python experiments/lora_finetuning/best_models/scripts/compare_best_models_to_excel.py --lora-run lora_3_epoch_zero_shot
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "experiments" / "lora_finetuning" / "shared" / "scripts"))

from compare_common import (  # noqa: E402
    METRICS,
    RUN_GPT_BASE,
    RUN_LORA_2_EPOCH_ZERO_SHOT,
    extract_row,
    load_metrics_summary,
    load_registry,
)
from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEY = "7B"
RIGHT_LABEL = "GPT-4o-mini"


def lora_run(registry: dict[str, Any], run_id: str) -> dict[str, Any]:
    runs = {run["run_id"]: run for run in registry[MODEL_KEY]["runs"]}
    return runs[run_id]


def best_rows(results_root: Path, registry: dict[str, Any], run_id: str) -> dict[str, tuple[list, list]]:
    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, run_id)
    lora_summary = load_metrics_summary(results_root / model_dir / run["folder"] / "metrics_summary.json")
    gpt_summary = load_metrics_summary(results_root / RUN_GPT_BASE / "metrics_summary.json")
    return {lang: (extract_row(lora_summary, lang), extract_row(gpt_summary, lang)) for lang in LANG_ORDER}


def write_sheet(ws: Worksheet, rows_by_lang: dict[str, tuple[list, list]], left_label: str) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 2

    lang_header = ws.cell(row=1, column=1, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    left_cell = ws.cell(row=1, column=data_start_col, value=left_label)
    apply_cell_style(left_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=data_start_col, end_row=1, end_column=data_start_col + cols_per_group - 1)

    right_col = data_start_col + cols_per_group
    right_cell = ws.cell(row=1, column=right_col, value=RIGHT_LABEL)
    apply_cell_style(right_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=right_col, end_row=1, end_column=right_col + cols_per_group - 1)

    for group_start in (data_start_col, right_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)

    for row_offset, lang in enumerate(LANG_ORDER, start=3):
        lang_cell = ws.cell(row=row_offset, column=1, value=lang)
        apply_cell_style(lang_cell)
        left_values, right_values = rows_by_lang[lang]

        for offset in range(cols_per_group):
            fills = rank_fills({"left": left_values[offset], "right": right_values[offset]})
            left_fill_font = fills.get("left")
            right_fill_font = fills.get("right")

            left_cell = ws.cell(row=row_offset, column=data_start_col + offset, value=left_values[offset])
            apply_cell_style(
                left_cell,
                fill=left_fill_font[0] if left_fill_font else None,
                font=left_fill_font[1] if left_fill_font else None,
            )
            right_cell = ws.cell(row=row_offset, column=right_col + offset, value=right_values[offset])
            apply_cell_style(
                right_cell,
                fill=right_fill_font[0] if right_fill_font else None,
                font=right_fill_font[1] if right_fill_font else None,
            )

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/lora_finetuning/shared/results"),
        help="Root directory containing gpt_base/ and Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/lora_finetuning/shared/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--lora-run",
        default=RUN_LORA_2_EPOCH_ZERO_SHOT,
        help="run_registry.json run_id (7B) to use as the 'best' LoRA config",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/lora_finetuning/best_models/report/best_models.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any], run_id: str) -> None:
    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, run_id)
    required = [
        results_root / model_dir / run["folder"] / "metrics_summary.json",
        results_root / RUN_GPT_BASE / "metrics_summary.json",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry, args.lora_run)

    model_dir = registry[MODEL_KEY]["model_dir"]
    run = lora_run(registry, args.lora_run)
    num_epochs = run["num_epochs"]
    left_label = f"qwen_lora_7b_zero_shot_{num_epochs}_epochs"

    rows_by_lang = best_rows(results_root, registry, args.lora_run)

    wb = Workbook()
    ws = wb.active
    ws.title = "best_models"
    write_sheet(ws, rows_by_lang, left_label)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 1 sheet ({len(LANG_ORDER)} languages, {model_dir}/{run['folder']} vs GPT) to {output_path}")


if __name__ == "__main__":
    main()
