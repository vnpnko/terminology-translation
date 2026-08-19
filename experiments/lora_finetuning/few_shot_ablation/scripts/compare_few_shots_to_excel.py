"""Compare zero-shot vs few-shot for GPT-4o-mini, Qwen base, and Qwen LoRA (1 epoch).

Writes one .xlsx file (default:
``experiments/lora_finetuning/few_shot_ablation/report/zero_shot_vs_few_shot_ablation.xlsx``) with 3 sheets:
``GPT-4o-mini``, ``Qwen2.5-3B``, ``Qwen2.5-7B``. The GPT sheet has one row per language pair;
the Qwen sheets stack a ``qwen_base`` block and a ``qwen_lora`` block (3 rows each), matching
the two differently-scoped few-shot experiments documented in ``report/README.md`` §3.4.1:

- **Baseline-level** (GPT, Qwen base): ``few_shot`` reads the LoRA experiment's own local
  baseline run (``<results-root>/gpt_base/`` or ``<results-root>/<model_dir>/qwen_base/``);
  ``zero_shot`` reads the separate, shared ``<baseline-results-root>/dev_v1/original/zero_shot/
  {gpt,qwen_3b,qwen_7b}/`` tree (not tracked in ``run_registry.json``).
- **LoRA-level** (``qwen_lora``): both columns are local, driven by ``run_registry.json``'s
  ``lora_1_epoch_zero_shot``/``lora_1_epoch_few_shot`` runs.

Each value cell is colored by the shared Good/Bad/Neutral convention (see
``src/analysis/excel_style.py``): green for the higher of its zero_shot/
few_shot pair, red for the lower, yellow if they're equal.

Usage::

    python experiments/lora_finetuning/few_shot_ablation/scripts/compare_few_shots_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEYS = ("3B", "7B")
BASELINE_DIRS = {"3B": "qwen_3b", "7B": "qwen_7b"}

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def lora_1_epoch_runs(registry: dict[str, Any], model_key: str) -> tuple[dict[str, Any], dict[str, Any]]:
    """Return (zero_shot_run, few_shot_run) for 1-epoch LoRA runs."""
    runs = {run["run_id"]: run for run in registry[model_key]["runs"]}
    return runs["lora_1_epoch_zero_shot"], runs["lora_1_epoch_few_shot"]


def gpt_rows(results_root: Path, baseline_results_root: Path) -> dict[str, tuple[list, list]]:
    zero_shot_summary = load_metrics_summary(
        baseline_results_root / "dev_v1" / "original" / "zero_shot" / "gpt" / "metrics_summary.json"
    )
    few_shot_summary = load_metrics_summary(results_root / "gpt_base" / "metrics_summary.json")
    return {
        lang: (extract_row(zero_shot_summary, lang), extract_row(few_shot_summary, lang)) for lang in LANG_ORDER
    }


def qwen_base_rows(
    results_root: Path, baseline_results_root: Path, model_dir: str, baseline_dir: str
) -> dict[str, tuple[list, list]]:
    zero_shot_summary = load_metrics_summary(
        baseline_results_root / "dev_v1" / "original" / "zero_shot" / baseline_dir / "metrics_summary.json"
    )
    few_shot_summary = load_metrics_summary(results_root / model_dir / "qwen_base" / "metrics_summary.json")
    return {
        lang: (extract_row(zero_shot_summary, lang), extract_row(few_shot_summary, lang)) for lang in LANG_ORDER
    }


def qwen_lora_rows(results_root: Path, registry: dict[str, Any], model_key: str, model_dir: str) -> dict[str, tuple[list, list]]:
    zero_shot_run, few_shot_run = lora_1_epoch_runs(registry, model_key)
    zero_shot_summary = load_metrics_summary(results_root / model_dir / zero_shot_run["folder"] / "metrics_summary.json")
    few_shot_summary = load_metrics_summary(results_root / model_dir / few_shot_run["folder"] / "metrics_summary.json")
    return {
        lang: (extract_row(zero_shot_summary, lang), extract_row(few_shot_summary, lang)) for lang in LANG_ORDER
    }


def write_group_header(ws: Worksheet, start_col: int, cols_per_group: int) -> None:
    zero_shot_cell = ws.cell(row=1, column=start_col, value="zero_shot")
    apply_cell_style(zero_shot_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + cols_per_group - 1)

    few_shot_col = start_col + cols_per_group
    few_shot_cell = ws.cell(row=1, column=few_shot_col, value="few_shot")
    apply_cell_style(few_shot_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=few_shot_col, end_row=1, end_column=few_shot_col + cols_per_group - 1)

    for group_start in (start_col, few_shot_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)


def write_data_row(ws: Worksheet, row: int, start_col: int, zero_shot_values: list, few_shot_values: list) -> None:
    cols_per_group = len(METRICS)
    for offset in range(cols_per_group):
        fills = rank_fills({"zero_shot": zero_shot_values[offset], "few_shot": few_shot_values[offset]})
        zero_fill_font = fills.get("zero_shot")
        few_fill_font = fills.get("few_shot")

        zero_cell = ws.cell(row=row, column=start_col + offset, value=zero_shot_values[offset])
        apply_cell_style(
            zero_cell,
            fill=zero_fill_font[0] if zero_fill_font else None,
            font=zero_fill_font[1] if zero_fill_font else None,
        )
        few_cell = ws.cell(row=row, column=start_col + cols_per_group + offset, value=few_shot_values[offset])
        apply_cell_style(
            few_cell,
            fill=few_fill_font[0] if few_fill_font else None,
            font=few_fill_font[1] if few_fill_font else None,
        )


def write_gpt_sheet(ws: Worksheet, rows_by_lang: dict[str, tuple[list, list]]) -> None:
    cols_per_group = len(METRICS)

    lang_header = ws.cell(row=1, column=1, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    write_group_header(ws, start_col=2, cols_per_group=cols_per_group)

    for row_offset, lang in enumerate(LANG_ORDER, start=3):
        lang_cell = ws.cell(row=row_offset, column=1, value=lang)
        apply_cell_style(lang_cell)
        zero_shot_values, few_shot_values = rows_by_lang[lang]
        write_data_row(ws, row_offset, start_col=2, zero_shot_values=zero_shot_values, few_shot_values=few_shot_values)

    autofit_columns(ws)


def write_qwen_sheet(
    ws: Worksheet,
    model_dir: str,
    base_rows_by_lang: dict[str, tuple[list, list]],
    lora_rows_by_lang: dict[str, tuple[list, list]],
) -> None:
    cols_per_group = len(METRICS)
    data_start_col = 3

    model_header = ws.cell(row=1, column=1, value=model_dir)
    apply_cell_style(model_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    lang_header = ws.cell(row=1, column=2, value="lang_pair")
    apply_cell_style(lang_header, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    write_group_header(ws, start_col=data_start_col, cols_per_group=cols_per_group)

    row = 3
    for block_label, rows_by_lang in (("qwen_base", base_rows_by_lang), ("qwen_lora", lora_rows_by_lang)):
        block_start_row = row
        for lang in LANG_ORDER:
            lang_cell = ws.cell(row=row, column=2, value=lang)
            apply_cell_style(lang_cell)
            zero_shot_values, few_shot_values = rows_by_lang[lang]
            write_data_row(ws, row, start_col=data_start_col, zero_shot_values=zero_shot_values, few_shot_values=few_shot_values)
            row += 1

        block_cell = ws.cell(row=block_start_row, column=1, value=block_label)
        apply_cell_style(block_cell)
        ws.merge_cells(start_row=block_start_row, start_column=1, end_row=row - 1, end_column=1)

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/lora_finetuning/results"),
        help="Root directory containing gpt_base/ and Qwen2.5-3B/, Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--baseline-results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/original/zero_shot/{gpt,qwen_3b,qwen_7b}/",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/lora_finetuning/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/lora_finetuning/few_shot_ablation/report/zero_shot_vs_few_shot_ablation.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def validate_required(results_root: Path, registry: dict[str, Any]) -> None:
    required = [results_root / "gpt_base" / "metrics_summary.json"]
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        required.append(results_root / model_dir / "qwen_base" / "metrics_summary.json")
        zero_shot_run, few_shot_run = lora_1_epoch_runs(registry, model_key)
        required.append(results_root / model_dir / zero_shot_run["folder"] / "metrics_summary.json")
        required.append(results_root / model_dir / few_shot_run["folder"] / "metrics_summary.json")

    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required metrics file(s): {missing}")


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    baseline_results_root = args.baseline_results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_required(results_root, registry)

    wb = Workbook()
    wb.remove(wb.active)

    gpt_ws = wb.create_sheet(title="GPT-4o-mini")
    write_gpt_sheet(gpt_ws, gpt_rows(results_root, baseline_results_root))

    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        baseline_dir = BASELINE_DIRS[model_key]
        base_rows = qwen_base_rows(results_root, baseline_results_root, model_dir, baseline_dir)
        lora_rows = qwen_lora_rows(results_root, registry, model_key, model_dir)
        ws = wb.create_sheet(title=model_dir)
        write_qwen_sheet(ws, model_dir, base_rows, lora_rows)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote 3 sheets (GPT-4o-mini, Qwen2.5-3B, Qwen2.5-7B) to {output_path}")


if __name__ == "__main__":
    main()
