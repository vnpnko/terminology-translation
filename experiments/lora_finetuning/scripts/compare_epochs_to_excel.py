"""Compare LoRA epoch counts (1/2/3) for each Qwen model size.

Writes one .xlsx file (default: ``experiments/lora_finetuning/report/epoch_ablation.xlsx``)
with one sheet per model size (``Qwen2.5-3B``, ``Qwen2.5-7B``). Each sheet has one row per
language pair and one column group per epoch count, with the non-baseline runs and their
epoch counts read from ``scripts/run_registry.json`` (not hardcoded). Reads
``metrics_summary.json`` from ``<results-root>/<model_dir>/<run_folder>/`` (``proper_term``
mode only).

Each value cell is colored by ranking that (language, metric) value **across the 3 epoch
counts** using the shared Good/Bad/Neutral convention (see ``src/analysis/excel_style.py``):
green = best epoch, red = worst epoch, the remaining (strictly middle) epoch is left unfilled.

Usage::

    python experiments/lora_finetuning/scripts/compare_epochs_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enes", "enru")
MODEL_KEYS = ("3B", "7B")
REQUIRED_EPOCHS = (1, 2, 3)

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Acc (%)", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Cons Macro Avg", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Cons Weighted Avg",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def epoch_runs_for_model(registry: dict[str, Any], model_key: str) -> list[tuple[int, dict[str, Any]]]:
    """Pick the zero-shot LoRA run for each epoch count, holding few-shot status constant."""
    runs = [run for run in registry[model_key]["runs"] if not run["is_baseline"] and not run["use_few_shot"]]
    by_epoch = {run["num_epochs"]: run for run in runs}

    missing = [epoch for epoch in REQUIRED_EPOCHS if epoch not in by_epoch]
    if missing:
        raise ValueError(
            f"run_registry.json model '{model_key}' is missing non-baseline, zero-shot runs "
            f"for epoch(s) {missing} (found: {sorted(by_epoch)})"
        )

    return [(epoch, by_epoch[epoch]) for epoch in REQUIRED_EPOCHS]


def load_metrics_summary(results_root: Path, model_dir: str, folder: str) -> dict[str, Any]:
    path = results_root / model_dir / folder / "metrics_summary.json"
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any], lang: str) -> list[float | None]:
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def validate_all_runs(results_root: Path, registry: dict[str, Any]) -> None:
    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        for _, run in epoch_runs_for_model(registry, model_key):
            summary_path = results_root / model_dir / run["folder"] / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(f"Missing metrics file for {model_key}/{run['folder']}: {summary_path}")


def build_model_rows(
    results_root: Path, registry: dict[str, Any], model_key: str
) -> dict[str, dict[int, list[float | None]]]:
    model_dir = registry[model_key]["model_dir"]
    epoch_runs = epoch_runs_for_model(registry, model_key)

    summaries = {
        epoch: load_metrics_summary(results_root, model_dir, run["folder"]) for epoch, run in epoch_runs
    }

    return {
        lang: {epoch: extract_row(summary, lang) for epoch, summary in summaries.items()} for lang in LANG_ORDER
    }


def epoch_label(num_epochs: int) -> str:
    return f"{num_epochs} epoch" if num_epochs == 1 else f"{num_epochs} epochs"


def write_sheet(ws: Worksheet, rows_by_lang: dict[str, dict[int, list[float | None]]]) -> None:
    cols_per_epoch = len(METRICS)

    header_cell = ws.cell(row=1, column=1, value="lang_pair")
    apply_cell_style(header_cell, fill=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    metric_start_col = 2
    for epoch_idx, num_epochs in enumerate(REQUIRED_EPOCHS):
        start_col = metric_start_col + epoch_idx * cols_per_epoch
        end_col = start_col + cols_per_epoch - 1

        group_cell = ws.cell(row=1, column=start_col, value=epoch_label(num_epochs))
        apply_cell_style(group_cell, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=title)
            apply_cell_style(sub_cell, fill=HEADER_FILL)

    for row_offset, lang in enumerate(LANG_ORDER, start=3):
        lang_cell = ws.cell(row=row_offset, column=1, value=lang)
        apply_cell_style(lang_cell)
        epoch_values = rows_by_lang[lang]

        for metric_idx in range(cols_per_epoch):
            values_by_epoch = {
                num_epochs: epoch_values[num_epochs][metric_idx] for num_epochs in REQUIRED_EPOCHS
            }
            fills = rank_fills(values_by_epoch)

            for epoch_idx, num_epochs in enumerate(REQUIRED_EPOCHS):
                start_col = metric_start_col + epoch_idx * cols_per_epoch
                cell = ws.cell(row=row_offset, column=start_col + metric_idx, value=values_by_epoch[num_epochs])
                fill_font = fills.get(num_epochs)
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                )

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("experiments/lora_finetuning/results"),
        help="Root directory containing Qwen2.5-3B/ and Qwen2.5-7B/ result folders",
    )
    parser.add_argument(
        "--registry",
        type=Path,
        default=Path("experiments/lora_finetuning/scripts/run_registry.json"),
        help="Path to run_registry.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/lora_finetuning/report/epoch_ablation.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    registry = load_registry(args.registry.resolve())
    validate_all_runs(results_root, registry)

    wb = Workbook()
    wb.remove(wb.active)

    for model_key in MODEL_KEYS:
        model_dir = registry[model_key]["model_dir"]
        rows_by_lang = build_model_rows(results_root, registry, model_key)
        ws = wb.create_sheet(title=model_dir)
        write_sheet(ws, rows_by_lang)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote {len(MODEL_KEYS)} sheets ({len(LANG_ORDER)} languages x {len(REQUIRED_EPOCHS)} epochs) to {output_path}")


if __name__ == "__main__":
    main()
