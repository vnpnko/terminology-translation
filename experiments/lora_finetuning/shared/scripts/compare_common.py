"""Shared helpers for the lora_finetuning sub-experiments' compare_*_to_excel.py scripts.

All 5 sub-experiment compare scripts (epoch_ablation, best_models, base_vs_lora,
few_shot_ablation, leakage_check) read metrics_summary.json files resolved via
run_registry.json and render them into a metrics-column-group Excel layout. This
module holds the logic that was previously duplicated verbatim across all 5.
"""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from shared.lib.analysis.excel_style import apply_cell_style

# run_registry.json's run_id vocabulary, referenced by name instead of as scattered literals.
RUN_GPT_BASE = "gpt_base"
RUN_BASE_FEW_SHOT = "base_few_shot"
RUN_LORA_1_EPOCH_FEW_SHOT = "lora_1_epoch_few_shot"
RUN_LORA_1_EPOCH_ZERO_SHOT = "lora_1_epoch_zero_shot"
RUN_LORA_2_EPOCH_ZERO_SHOT = "lora_2_epoch_zero_shot"
RUN_LORA_3_EPOCH_ZERO_SHOT = "lora_3_epoch_zero_shot"

METRICS = (
    ("bleu", "bleu", "BLEU", 2),
    ("chrf", "chrf", "chrF", 2),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %", 2),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency", 4),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
        4,
    ),
)


def load_registry(registry_path: Path) -> dict[str, Any]:
    with registry_path.open(encoding="utf-8") as f:
        return json.load(f)


def load_metrics_summary(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_row(summary: dict[str, Any] | None, lang: str) -> list[float | None]:
    if summary is None:
        return [None] * len(METRICS)
    metrics = summary["languages"][lang]["modes"]["proper_term"]["metrics"]
    values = []
    for _, spec, _, decimals in METRICS:
        value = extract_metric(metrics, spec)
        values.append(round(value, decimals) if value is not None else None)
    return values


def write_group_header(
    ws: Worksheet,
    start_col: int,
    cols_per_group: int,
    left_label: str,
    right_label: str,
    fill: PatternFill | None = None,
) -> None:
    left_cell = ws.cell(row=1, column=start_col, value=left_label)
    apply_cell_style(left_cell, fill=fill)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + cols_per_group - 1)

    right_col = start_col + cols_per_group
    right_cell = ws.cell(row=1, column=right_col, value=right_label)
    apply_cell_style(right_cell, fill=fill)
    ws.merge_cells(start_row=1, start_column=right_col, end_row=1, end_column=right_col + cols_per_group - 1)

    for group_start in (start_col, right_col):
        for offset, (_, _, title, _) in enumerate(METRICS):
            sub_cell = ws.cell(row=2, column=group_start + offset, value=title)
            apply_cell_style(sub_cell, fill=fill)
