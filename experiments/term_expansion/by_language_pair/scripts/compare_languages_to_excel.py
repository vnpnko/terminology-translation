"""Compare ende, enru, and enes language pairs across all baseline models and modes.

Writes one styled .xlsx file (default:
``experiments/term_expansion/by_language_pair/report/language_comparison.xlsx``)
with two sheets (``dev_v1`` -- from ``dev_v1/original/few_shot`` --, and
``dev_v2``), rows grouped by mode (``no_term``/``proper_term``/``random_term``)
then model. Reads ``metrics_summary.json`` from
``<results-root>/<variant>/{gpt,qwen_3b,qwen_7b}/``.

Each value cell (ende/enru/enes) is colored by ranking it against the other
two languages in the same row, using the shared Good/Bad/Neutral convention
(see ``src/analysis/excel_style.py``): green = best language, red = worst,
a genuine tie is neutral, and a strictly-middle value is left unfilled.

The ``expand``/``cleaned``/``dictionary`` term-list variants are covered by
``proper_term_across_languages.xlsx`` instead (see
``compare_proper_term_across_languages_to_excel.py``) -- they were dropped
from here since ``no_term``/``random_term`` don't exist for those variants
(only ``original/few_shot`` has all three modes), making their sheets here
pure duplicates of that other workbook's ``proper_term`` rows.

Usage::

    python experiments/term_expansion/by_language_pair/scripts/compare_languages_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from itertools import groupby
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
BASELINE_LABELS = {
    "gpt": "GPT",
    "qwen_3b": "Qwen 3B",
    "qwen_7b": "Qwen 7B",
}

DATASET_VARIANTS = (
    ("dev_v1", Path("dev_v1/original/few_shot"), BASELINE_DIRS),
    ("dev_v2", Path("dev_v2"), BASELINE_DIRS),
)

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict[str, Any]) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def validate_all_variants(results_root: Path) -> None:
    missing = [
        str(results_root / variant_path / baseline_dir / "metrics_summary.json")
        for _, variant_path, baseline_dirs in DATASET_VARIANTS
        for baseline_dir in baseline_dirs
        if not (results_root / variant_path / baseline_dir / "metrics_summary.json").exists()
    ]
    if missing:
        raise FileNotFoundError(f"Missing metrics file(s): {missing}")


def build_comparison(dataset_dir: Path, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}

    for baseline_dir in baseline_dirs:
        summary_path = dataset_dir / baseline_dir / "metrics_summary.json"
        if not summary_path.exists():
            raise FileNotFoundError(f"Missing metrics file: {summary_path}")
        summaries[baseline_dir] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for baseline_dir in baseline_dirs:
            lang_metrics = {
                lang: summaries[baseline_dir].get((lang, mode)) for lang in LANG_ORDER
            }
            if all(metrics is None for metrics in lang_metrics.values()):
                continue

            row: dict[str, object] = {"mode": mode, "model": BASELINE_LABELS[baseline_dir]}

            for column, _, _ in METRICS:
                for lang in LANG_ORDER:
                    metrics = lang_metrics[lang]
                    value = metrics.get(column) if metrics else None
                    row[f"{lang}_{column}"] = value

            rows.append(row)

    columns = ["mode", "model"]
    for column, _, _ in METRICS:
        columns.extend([f"{lang}_{column}" for lang in LANG_ORDER])

    return pd.DataFrame(rows, columns=columns)


def write_sheet(ws, df: pd.DataFrame, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> None:
    fixed_headers = ("mode", "model")
    value_subheaders = LANG_ORDER
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["model"] == BASELINE_LABELS[baseline_dirs[-1]]

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["model"] == BASELINE_LABELS[baseline_dirs[0]] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        model_cell = ws.cell(row=row_offset, column=2, value=record["model"])
        apply_cell_style(model_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            row_values = {lang: record[f"{lang}_{column}"] for lang in LANG_ORDER}
            fills = rank_fills(row_values)

            for offset, lang in enumerate(LANG_ORDER):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=row_values[lang])
                fill_font = fills.get(lang)
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    row_offset = 3
    for _mode, group in groupby(records, key=lambda r: r["mode"]):
        group_len = sum(1 for _ in group)
        end_row = row_offset + group_len - 1
        if end_row > row_offset:
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row_offset = end_row + 1

    autofit_columns(ws)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1/{original,expand,cleaned} and dev_v2 result folders",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("experiments/term_expansion/by_language_pair/report/language_comparison.xlsx"),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    validate_all_variants(results_root)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_title, variant_path, baseline_dirs in DATASET_VARIANTS:
        df = build_comparison(results_root / variant_path, baseline_dirs)
        ws = wb.create_sheet(title=sheet_title)
        write_sheet(ws, df, baseline_dirs)

    wb.save(output_path)

    variant_names = ", ".join(name for name, _, _ in DATASET_VARIANTS)
    print(f"Wrote {len(DATASET_VARIANTS)} sheets ({variant_names}) to {output_path}")


if __name__ == "__main__":
    main()
