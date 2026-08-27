"""Compare GPT/Qwen 3B/Qwen 7B and ende/enru/enes across term_expansion result variants.

Supports two modes via ``--mode``:

- ``--mode all`` (default) -- ``no_term``/``proper_term``/``random_term`` across
  ``dev_v1`` (``original/few_shot``) and ``dev_v2``. Writes two styled .xlsx files:

  - ``report/model_comparison.xlsx`` -- rows grouped by mode then language,
    columns per model (GPT/Qwen 3B/Qwen 7B). Each value cell is colored by
    ranking it against the other two models in the same row.
  - ``report/language_comparison.xlsx`` -- rows grouped by mode then model,
    columns per language (ende/enru/enes). Each value cell is colored by
    ranking it against the other two languages in the same row.

  Both use the shared Good/Bad/Neutral convention (see
  ``shared/lib/analysis/excel_style.py``): green = best, red = worst, a genuine
  tie is neutral, and a strictly-middle value is left unfilled. Each has two
  sheets, ``dev_v1`` (from ``dev_v1/original/few_shot``) and ``dev_v2``. Reads
  ``metrics_summary.json`` from ``<results-root>/<variant>/{gpt,qwen_3b,qwen_7b}/``.

  The ``expand``/``cleaned``/``dictionary`` term-list variants are covered by
  ``--mode proper_term`` instead -- they were dropped from here since
  ``no_term``/``random_term`` don't exist for those variants (only
  ``original/few_shot`` has all three modes), making their sheets here pure
  duplicates of that other mode's ``proper_term`` rows.

- ``--mode proper_term`` -- ``proper_term`` mode only, across the 4 dev_v1
  term-list variants (``original``, ``expand``, ``cleaned``, ``dictionary``).
  Writes two styled .xlsx files, each with 12 data rows (4 variants x 3
  languages/models):

  - ``report/proper_term_across_models.xlsx`` -- 3 language rows per variant,
    columns per model (GPT/Qwen 3B/Qwen 7B).
  - ``report/proper_term_across_languages.xlsx`` -- 3 model rows per variant,
    columns per language (ende/enru/enes).

  Reads ``metrics_summary.json`` from ``<variant_dir>/{gpt,qwen_3b,qwen_7b}/``.
  Each value cell is colored by ranking that (model, language) combination's
  value **across the 4 variants** (not across models/languages), using the
  shared Good/Bad/Neutral convention: green = best variant, red = worst
  variant, any value neither best nor worst is left unfilled.

  Note: ``shared/results/dev_v1/original/`` has no ``gpt``/``qwen_3b``/``qwen_7b``
  subfolders directly — it's nested under ``zero_shot/`` or ``few_shot/``
  (see ``report/README.md`` §3.4.1). This mode defaults to ``few_shot``;
  override with ``--original`` if you want the ``zero_shot`` variant instead.

Usage::

    python experiments/term_expansion/by_language_pair/shared/scripts/compare_by_model_and_language.py
    python experiments/term_expansion/by_language_pair/shared/scripts/compare_by_model_and_language.py --mode proper_term
    python experiments/term_expansion/by_language_pair/shared/scripts/compare_by_model_and_language.py --mode proper_term --original shared/results/dev_v1/original/zero_shot
"""

from __future__ import annotations

import argparse
import sys
from itertools import groupby
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[5]
sys.path.insert(0, str(PROJECT_ROOT))

from shared.lib.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    rank_fills,
)
from shared.lib.analysis.metrics_loader import (  # noqa: E402
    BASELINE_DIRS,
    BASELINE_LABELS,
    DEFAULT_MODE,
    LANG_ORDER,
    METRICS,
    MODE_ORDER,
    extract_metric,
    load_summary,
)

DATASET_VARIANTS = (
    ("dev_v1", Path("dev_v1/original/few_shot"), BASELINE_DIRS),
    ("dev_v2", Path("dev_v2"), BASELINE_DIRS),
)

VARIANT_ORDER = ("original", "expand", "cleaned", "dictionary")
PROPER_TERM_MODE = DEFAULT_MODE  # "proper_term"


# ---------------------------------------------------------------------------
# --mode all
# ---------------------------------------------------------------------------


def extract_mode_metrics(summary: dict) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def validate_all_variants(results_root: Path) -> None:
    missing = [
        str(results_root / variant_path / baseline_dir / "metrics_summary.json")
        for _, variant_path, baseline_dirs in DATASET_VARIANTS
        for baseline_dir in baseline_dirs
        if not (results_root / variant_path / baseline_dir / "metrics_summary.json").exists()
    ]
    if missing:
        raise FileNotFoundError(f"Missing metrics file(s): {missing}")


def _load_summaries(
    dataset_dir: Path, baseline_dirs: tuple[str, ...]
) -> dict[str, dict[tuple[str, str], dict[str, float | None]]]:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}
    for baseline_dir in baseline_dirs:
        summary_path = dataset_dir / baseline_dir / "metrics_summary.json"
        if not summary_path.exists():
            raise FileNotFoundError(f"Missing metrics file: {summary_path}")
        summaries[baseline_dir] = extract_mode_metrics(load_summary(summary_path))
    return summaries


def build_comparison_by_model(dataset_dir: Path, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> pd.DataFrame:
    summaries = _load_summaries(dataset_dir, baseline_dirs)
    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for lang in LANG_ORDER:
            baseline_metrics = {
                baseline_dir: summaries[baseline_dir].get((lang, mode))
                for baseline_dir in baseline_dirs
            }
            if all(metrics is None for metrics in baseline_metrics.values()):
                continue

            row: dict[str, object] = {"mode": mode, "language": lang}
            for column, _, _ in METRICS:
                for baseline_dir in baseline_dirs:
                    metrics = baseline_metrics[baseline_dir]
                    row[f"{baseline_dir}_{column}"] = metrics.get(column) if metrics else None
            rows.append(row)

    columns = ["mode", "language"]
    for column, _, _ in METRICS:
        columns.extend([f"{baseline_dir}_{column}" for baseline_dir in baseline_dirs])
    return pd.DataFrame(rows, columns=columns)


def build_comparison_by_language(dataset_dir: Path, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> pd.DataFrame:
    summaries = _load_summaries(dataset_dir, baseline_dirs)
    rows: list[dict[str, object]] = []

    for mode in MODE_ORDER:
        for baseline_dir in baseline_dirs:
            lang_metrics = {lang: summaries[baseline_dir].get((lang, mode)) for lang in LANG_ORDER}
            if all(metrics is None for metrics in lang_metrics.values()):
                continue

            row: dict[str, object] = {"mode": mode, "model": BASELINE_LABELS[baseline_dir]}
            for column, _, _ in METRICS:
                for lang in LANG_ORDER:
                    metrics = lang_metrics[lang]
                    row[f"{lang}_{column}"] = metrics.get(column) if metrics else None
            rows.append(row)

    columns = ["mode", "model"]
    for column, _, _ in METRICS:
        columns.extend([f"{lang}_{column}" for lang in LANG_ORDER])
    return pd.DataFrame(rows, columns=columns)


def write_sheet_by_model(ws, df: pd.DataFrame, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> None:
    fixed_headers = ("mode", "language")
    value_subheaders = tuple(BASELINE_LABELS[baseline_dir] for baseline_dir in baseline_dirs)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["language"] == "enes"

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["language"] == LANG_ORDER[0] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        apply_cell_style(lang_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            row_values = {
                baseline_dir: record[f"{baseline_dir}_{column}"] for baseline_dir in baseline_dirs
            }
            fills = rank_fills(row_values)

            for offset, baseline_dir in enumerate(baseline_dirs):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=row_values[baseline_dir])
                fill_font = fills.get(baseline_dir)
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    row_offset = 3
    for _mode, group in groupby(records, key=lambda r: r["mode"]):
        group_len = sum(1 for _ in group)
        end_row = row_offset + group_len - 1
        if end_row > row_offset:
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row_offset = end_row + 1

    autofit_columns(ws)


def write_sheet_by_language(ws, df: pd.DataFrame, baseline_dirs: tuple[str, ...] = BASELINE_DIRS) -> None:
    fixed_headers = ("mode", "model")
    value_subheaders = LANG_ORDER
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["model"] == BASELINE_LABELS[baseline_dirs[-1]]

        mode_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["mode"] if record["model"] == BASELINE_LABELS[baseline_dirs[0]] else None,
        )
        apply_cell_style(mode_cell, thick_bottom=thick_bottom)

        model_cell = ws.cell(row=row_offset, column=2, value=record["model"])
        apply_cell_style(model_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            row_values = {lang: record[f"{lang}_{column}"] for lang in LANG_ORDER}
            fills = rank_fills(row_values)

            for offset, lang in enumerate(LANG_ORDER):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=row_values[lang])
                fill_font = fills.get(lang)
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                    thick_bottom=thick_bottom,
                )

    row_offset = 3
    for _mode, group in groupby(records, key=lambda r: r["mode"]):
        group_len = sum(1 for _ in group)
        end_row = row_offset + group_len - 1
        if end_row > row_offset:
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row_offset = end_row + 1

    autofit_columns(ws)


def _run_all_mode(args: argparse.Namespace) -> None:
    results_root = args.results_root.resolve()
    validate_all_variants(results_root)

    model_output_path = args.model_output.resolve()
    model_output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_model = Workbook()
    wb_model.remove(wb_model.active)
    for sheet_title, variant_path, baseline_dirs in DATASET_VARIANTS:
        df = build_comparison_by_model(results_root / variant_path, baseline_dirs)
        ws = wb_model.create_sheet(title=sheet_title)
        write_sheet_by_model(ws, df, baseline_dirs)
    wb_model.save(model_output_path)

    language_output_path = args.language_output.resolve()
    language_output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_lang = Workbook()
    wb_lang.remove(wb_lang.active)
    for sheet_title, variant_path, baseline_dirs in DATASET_VARIANTS:
        df = build_comparison_by_language(results_root / variant_path, baseline_dirs)
        ws = wb_lang.create_sheet(title=sheet_title)
        write_sheet_by_language(ws, df, baseline_dirs)
    wb_lang.save(language_output_path)

    variant_names = ", ".join(name for name, _, _ in DATASET_VARIANTS)
    print(f"Wrote {len(DATASET_VARIANTS)} sheets ({variant_names}) to {model_output_path}")
    print(f"Wrote {len(DATASET_VARIANTS)} sheets ({variant_names}) to {language_output_path}")


# ---------------------------------------------------------------------------
# --mode proper_term
# ---------------------------------------------------------------------------


def extract_proper_term_metrics(summary: dict) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(PROPER_TERM_MODE)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {column: extract_metric(metrics, spec) for column, spec, _ in METRICS}
    return by_lang


def _load_proper_term_summaries(
    variant_dirs: dict[str, Path]
) -> dict[tuple[str, str], dict[str, dict[str, float | None]]]:
    summaries: dict[tuple[str, str], dict[str, dict[str, float | None]]] = {}
    for variant, variant_dir in variant_dirs.items():
        for baseline_dir in BASELINE_DIRS:
            summary_path = variant_dir / baseline_dir / "metrics_summary.json"
            if not summary_path.exists():
                raise FileNotFoundError(f"Missing metrics file: {summary_path}")
            summaries[(variant, baseline_dir)] = extract_proper_term_metrics(load_summary(summary_path))
    return summaries


def build_proper_term_comparison_by_model(variant_dirs: dict[str, Path]) -> dict[tuple[str, str], dict[str, object]]:
    summaries = _load_proper_term_summaries(variant_dirs)
    rows: dict[tuple[str, str], dict[str, object]] = {}
    for variant in VARIANT_ORDER:
        for lang in LANG_ORDER:
            baseline_metrics = {
                baseline_dir: summaries[(variant, baseline_dir)].get(lang) for baseline_dir in BASELINE_DIRS
            }
            if all(metrics is None for metrics in baseline_metrics.values()):
                continue

            row: dict[str, object] = {"data": variant, "language": lang}
            for column, _, _ in METRICS:
                for baseline_dir in BASELINE_DIRS:
                    metrics = baseline_metrics[baseline_dir]
                    row[f"{baseline_dir}_{column}"] = metrics.get(column) if metrics else None
            rows[(variant, lang)] = row
    return rows


def build_proper_term_comparison_by_language(
    variant_dirs: dict[str, Path]
) -> dict[tuple[str, str], dict[str, object]]:
    summaries = _load_proper_term_summaries(variant_dirs)
    rows: dict[tuple[str, str], dict[str, object]] = {}
    for variant in VARIANT_ORDER:
        for baseline_dir in BASELINE_DIRS:
            lang_metrics = summaries[(variant, baseline_dir)]
            if not lang_metrics:
                continue

            row: dict[str, object] = {"data": variant, "model": BASELINE_LABELS[baseline_dir]}
            for column, _, _ in METRICS:
                for lang in LANG_ORDER:
                    value = lang_metrics.get(lang, {}).get(column)
                    row[f"{lang}_{column}"] = value
            rows[(variant, baseline_dir)] = row
    return rows


def _proper_term_variant_rank_fill(
    rows: dict[tuple[str, str], dict[str, object]], sub_key: str, key: str, variant: str
):
    """Rank a value against its counterpart in the other 3 variants."""
    values = {
        v: rows[(v, sub_key)][key]
        for v in VARIANT_ORDER
        if (v, sub_key) in rows and rows[(v, sub_key)][key] is not None
    }
    return rank_fills(values).get(variant)


def write_proper_term_excel_by_model(rows: dict[tuple[str, str], dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "proper_term"

    fixed_headers = ("data", "language")
    value_subheaders = tuple(BASELINE_LABELS[baseline_dir] for baseline_dir in BASELINE_DIRS)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    row_offset = 3
    for variant in VARIANT_ORDER:
        for lang_idx, lang in enumerate(LANG_ORDER):
            record = rows.get((variant, lang))
            if record is None:
                continue

            data_cell = ws.cell(row=row_offset, column=1, value=variant if lang_idx == 0 else None)
            apply_cell_style(data_cell)

            lang_cell = ws.cell(row=row_offset, column=2, value=lang)
            apply_cell_style(lang_cell)

            for metric_idx, (column, _, _) in enumerate(METRICS):
                start_col = metric_start_col + metric_idx * cols_per_metric
                baseline_key_pairs = [(f"{baseline_dir}_{column}", baseline_dir) for baseline_dir in BASELINE_DIRS]

                for offset, (key, _baseline_dir) in enumerate(baseline_key_pairs):
                    cell = ws.cell(row=row_offset, column=start_col + offset, value=record[key])
                    fill_font = _proper_term_variant_rank_fill(rows, lang, key, variant)
                    apply_cell_style(
                        cell,
                        fill=fill_font[0] if fill_font else None,
                        font=fill_font[1] if fill_font else None,
                    )

            row_offset += 1

    for variant_idx in range(len(VARIANT_ORDER)):
        start_row = 3 + variant_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    autofit_columns(ws)
    wb.save(output_path)


def write_proper_term_excel_by_language(rows: dict[tuple[str, str], dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "proper_term"

    fixed_headers = ("data", "model")
    value_subheaders = LANG_ORDER
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    row_offset = 3
    for variant in VARIANT_ORDER:
        for baseline_idx, baseline_dir in enumerate(BASELINE_DIRS):
            record = rows.get((variant, baseline_dir))
            if record is None:
                continue

            data_cell = ws.cell(row=row_offset, column=1, value=variant if baseline_idx == 0 else None)
            apply_cell_style(data_cell)

            model_cell = ws.cell(row=row_offset, column=2, value=BASELINE_LABELS[baseline_dir])
            apply_cell_style(model_cell)

            for metric_idx, (column, _, _) in enumerate(METRICS):
                start_col = metric_start_col + metric_idx * cols_per_metric
                lang_keys = [f"{lang}_{column}" for lang in LANG_ORDER]

                for offset, key in enumerate(lang_keys):
                    cell = ws.cell(row=row_offset, column=start_col + offset, value=record[key])
                    fill_font = _proper_term_variant_rank_fill(rows, baseline_dir, key, variant)
                    apply_cell_style(
                        cell,
                        fill=fill_font[0] if fill_font else None,
                        font=fill_font[1] if fill_font else None,
                    )

            row_offset += 1

    for variant_idx in range(len(VARIANT_ORDER)):
        start_row = 3 + variant_idx * len(BASELINE_DIRS)
        end_row = start_row + len(BASELINE_DIRS) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    autofit_columns(ws)
    wb.save(output_path)


def _run_proper_term_mode(args: argparse.Namespace) -> None:
    results_root = args.results_root.resolve()
    original_dir = args.original.resolve() if args.original else results_root / "dev_v1" / "original" / "few_shot"
    variant_dirs = {
        "original": original_dir,
        "expand": results_root / "dev_v1" / "expand",
        "cleaned": results_root / "dev_v1" / "cleaned",
        "dictionary": results_root / "dev_v1" / "dictionary",
    }

    model_rows = build_proper_term_comparison_by_model(variant_dirs)
    model_output_path = args.model_output.resolve()
    model_output_path.parent.mkdir(parents=True, exist_ok=True)
    write_proper_term_excel_by_model(model_rows, model_output_path)

    language_rows = build_proper_term_comparison_by_language(variant_dirs)
    language_output_path = args.language_output.resolve()
    language_output_path.parent.mkdir(parents=True, exist_ok=True)
    write_proper_term_excel_by_language(language_rows, language_output_path)

    print(
        f"Wrote {len(model_rows)} rows "
        f"({len(VARIANT_ORDER)} variants x {len(LANG_ORDER)} languages) to {model_output_path}"
    )
    print(
        f"Wrote {len(language_rows)} rows "
        f"({len(VARIANT_ORDER)} variants x {len(BASELINE_DIRS)} models) to {language_output_path}"
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

MODE_DEFAULTS = {
    "all": {
        "model_output": Path("experiments/term_expansion/by_model/report/model_comparison.xlsx"),
        "language_output": Path("experiments/term_expansion/by_language_pair/report/language_comparison.xlsx"),
    },
    "proper_term": {
        "model_output": Path("experiments/term_expansion/by_model/report/proper_term_across_models.xlsx"),
        "language_output": Path(
            "experiments/term_expansion/by_language_pair/report/proper_term_across_languages.xlsx"
        ),
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--mode",
        choices=("all", "proper_term"),
        default="all",
        help="'all' compares no_term/proper_term/random_term across dev_v1/dev_v2 (default); "
        "'proper_term' compares proper_term mode across the 4 dev_v1 term-list variants",
    )
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("shared/results"),
        help="Root directory containing dev_v1/{original,expand,cleaned,dictionary} and dev_v2 result folders",
    )
    parser.add_argument(
        "--original",
        type=Path,
        default=None,
        help=(
            "--mode proper_term only: path to the dev_v1/original results (default: "
            "<results-root>/dev_v1/original/few_shot — "
            "override with <results-root>/dev_v1/original/zero_shot if preferred)"
        ),
    )
    parser.add_argument(
        "--model-output",
        type=Path,
        default=None,
        help="Output .xlsx path for the by-model workbook (default depends on --mode)",
    )
    parser.add_argument(
        "--language-output",
        type=Path,
        default=None,
        help="Output .xlsx path for the by-language workbook (default depends on --mode)",
    )
    args = parser.parse_args()

    defaults = MODE_DEFAULTS[args.mode]
    if args.model_output is None:
        args.model_output = defaults["model_output"]
    if args.language_output is None:
        args.language_output = defaults["language_output"]
    return args


def main() -> None:
    args = parse_args()
    if args.mode == "all":
        _run_all_mode(args)
    else:
        _run_proper_term_mode(args)


if __name__ == "__main__":
    main()
