"""Compare original, expand, cleaned, and dictionary term-list variants for GPT, proper_term mode.

Writes one styled .xlsx file (default:
``experiments/term_expansion/dataset_comparison/report/dev_v1_proper_term_across_expansion_modes.xlsx``)
with 3 data rows (ende, enru, enes), all under a single merged ``proper_term``
mode cell. Each metric block has 4 variant columns (original / expand /
cleaned / dictionary) plus a ``best`` column. GPT only — dictionary results
don't exist for Qwen 3B/7B (``results/dev_v1/dictionary/`` has only
``gpt/``). Reads ``metrics_summary.json`` from ``<variant_dir>/gpt/``.

Note: ``results/dev_v1/original/`` has no ``gpt/`` directly — it's nested
under ``zero_shot/`` or ``few_shot/`` (see ``report/README.md``
§3.4.1). This script defaults to ``few_shot``; override with
``--original`` if you want the ``zero_shot`` variant instead.

Usage::

    python experiments/term_expansion/dataset_comparison/scripts/compare_proper_term_across_expansion_modes_to_excel.py
    python experiments/term_expansion/dataset_comparison/scripts/compare_proper_term_across_expansion_modes_to_excel.py --original results/dev_v1/original/zero_shot
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

from src.analysis.excel_style import (  # noqa: E402
    HEADER_FILL,
    apply_cell_style,
    autofit_columns,
    best_label,
    label_fill,
)

LANG_ORDER = ("ende", "enru", "enes")
VARIANT_ORDER = ("original", "expand", "cleaned", "dictionary")
MODE = "proper_term"

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_proper_term_metrics(summary: dict[str, Any]) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(MODE)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {column: extract_metric(metrics, spec) for column, spec, _ in METRICS}
    return by_lang


def build_comparison(variant_dirs: dict[str, Path]) -> list[dict[str, object]]:
    summaries: dict[str, dict[str, dict[str, float | None]]] = {}

    for variant, variant_dir in variant_dirs.items():
        summary_path = variant_dir / "gpt" / "metrics_summary.json"
        if not summary_path.exists():
            raise FileNotFoundError(f"Missing metrics file: {summary_path}")
        summaries[variant] = extract_proper_term_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []
    for lang in LANG_ORDER:
        variant_metrics = {variant: summaries[variant].get(lang) for variant in VARIANT_ORDER}
        if all(metrics is None for metrics in variant_metrics.values()):
            continue

        row: dict[str, object] = {"mode": MODE, "language": lang}

        for column, _, _ in METRICS:
            labeled_values: dict[str, float | None] = {}
            for variant in VARIANT_ORDER:
                metrics = variant_metrics[variant]
                value = metrics.get(column) if metrics else None
                row[f"{variant}_{column}"] = value
                labeled_values[variant] = value

            row[f"best_{column}"] = best_label(labeled_values)

        rows.append(row)

    return rows


def write_styled_excel(rows: list[dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "proper_term"

    fixed_headers = ("mode", "language")
    value_subheaders = VARIANT_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    for row_offset, record in enumerate(rows, start=3):
        is_first_lang = record["language"] == LANG_ORDER[0]

        mode_cell = ws.cell(row=row_offset, column=1, value=MODE if is_first_lang else None)
        apply_cell_style(mode_cell)

        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        apply_cell_style(lang_cell)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{variant}_{column}"] for variant in VARIANT_ORDER]
            values.append(record[f"best_{column}"])

            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill_font = label_fill(value) if offset == cols_per_metric - 1 else None
                apply_cell_style(
                    cell,
                    fill=fill_font[0] if fill_font else None,
                    font=fill_font[1] if fill_font else None,
                )

    if rows:
        ws.merge_cells(start_row=3, start_column=1, end_row=2 + len(rows), end_column=1)
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")

    autofit_columns(ws)

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--results-root",
        type=Path,
        default=Path("results"),
        help="Root directory containing dev_v1 result folders",
    )
    parser.add_argument(
        "--original",
        type=Path,
        default=None,
        help=(
            "Path to the dev_v1/original results (default: "
            "<results-root>/dev_v1/original/few_shot — "
            "override with <results-root>/dev_v1/original/zero_shot if preferred)"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "experiments/term_expansion/dataset_comparison/report/dev_v1_proper_term_across_expansion_modes.xlsx"
        ),
        help="Output .xlsx path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    results_root = args.results_root.resolve()
    variant_dirs = {
        "original": (args.original.resolve() if args.original else results_root / "dev_v1" / "original" / "few_shot"),
        "expand": results_root / "dev_v1" / "expand",
        "cleaned": results_root / "dev_v1" / "cleaned",
        "dictionary": results_root / "dev_v1" / "dictionary",
    }

    rows = build_comparison(variant_dirs)
    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_styled_excel(rows, output_path)

    print(f"Wrote {len(rows)} rows ({len(LANG_ORDER)} languages) to {output_path}")


if __name__ == "__main__":
    main()
