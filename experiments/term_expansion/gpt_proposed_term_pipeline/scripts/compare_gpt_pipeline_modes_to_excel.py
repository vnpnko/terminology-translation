"""Compare GPT `proper_term` (oracle dictionary) vs `gpt_proposed_term` (self-proposed) modes.

Both runs are zero-shot GPT-4o-mini on dev_v1/original, so they are directly
comparable. Reads:
    shared/results/dev_v1/original/zero_shot/gpt/metrics_summary.json (mode: proper_term)
    experiments/term_expansion/gpt_proposed_term_pipeline/results/metrics_summary.json
        (mode: gpt_proposed_term, produced by compute_gpt_proposed_metrics.py)

Writes one styled .xlsx with two sheets: a `modes` metric comparison (BLEU/chrF/
terminology accuracy/consistency, oracle_diagnostics only apply to gpt_proposed_term so
carried on that side) and an `oracle_diagnostics` sheet showing how well GPT's own
extraction/proposal matched the oracle dictionary.

Usage::

    python experiments/term_expansion/gpt_proposed_term_pipeline/scripts/compare_gpt_pipeline_modes_to_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

EXPERIMENT_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

from shared.lib.analysis.excel_style import HEADER_FILL, autofit_columns, best_label, label_fill  # noqa: E402

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("proper_term", "gpt_proposed_term")

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)

THIN = Side(style="thin", color="000000")
THICK = Side(style="medium", color="000000")


def load_summary(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict[str, Any], spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def mode_metrics_by_lang(summary: dict[str, Any], mode: str) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(mode)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {column: extract_metric(metrics, spec) for column, spec, _ in METRICS}
    return by_lang


def oracle_diagnostics_by_lang(summary: dict[str, Any], mode: str) -> dict[str, dict[str, Any]]:
    by_lang: dict[str, dict[str, Any]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(mode)
        if not mode_data:
            continue
        diag = mode_data.get("metrics", {}).get("oracle_diagnostics")
        if diag:
            by_lang[lang] = diag
    return by_lang


def build_comparison(oracle_summary: dict[str, Any], pipeline_summary: dict[str, Any]) -> pd.DataFrame:
    by_mode = {
        "proper_term": mode_metrics_by_lang(oracle_summary, "proper_term"),
        "gpt_proposed_term": mode_metrics_by_lang(pipeline_summary, "gpt_proposed_term"),
    }

    rows: list[dict[str, object]] = []
    for lang in LANG_ORDER:
        mode_metrics = {mode: by_mode[mode].get(lang) for mode in MODE_ORDER}
        if all(m is None for m in mode_metrics.values()):
            continue
        row: dict[str, object] = {"language": lang}
        for column, _, _ in METRICS:
            labeled: dict[str, float | None] = {}
            for mode in MODE_ORDER:
                metrics = mode_metrics[mode]
                value = metrics.get(column) if metrics else None
                row[f"{mode}_{column}"] = value
                labeled[mode] = value
            row[f"best_{column}"] = best_label(labeled)
        rows.append(row)

    columns = ["language"]
    for column, _, _ in METRICS:
        columns.extend([f"{m}_{column}" for m in MODE_ORDER])
        columns.append(f"best_{column}")
    return pd.DataFrame(rows, columns=columns)


def build_oracle_diagnostics_table(pipeline_summary: dict[str, Any]) -> pd.DataFrame:
    diag_by_lang = oracle_diagnostics_by_lang(pipeline_summary, "gpt_proposed_term")
    rows = []
    for lang in LANG_ORDER:
        diag = diag_by_lang.get(lang)
        if not diag:
            continue
        rows.append(
            {
                "language": lang,
                "extraction_overlap_pct": diag.get("extraction_overlap_pct"),
                "extraction_hits": diag.get("extraction_hits"),
                "extraction_total": diag.get("extraction_total"),
                "proposal_match_pct": diag.get("proposal_match_pct"),
                "proposal_hits": diag.get("proposal_hits"),
                "proposal_total": diag.get("proposal_total"),
            }
        )
    return pd.DataFrame(rows)


def write_modes_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "modes"

    fixed_headers = ("language",)
    value_subheaders = MODE_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)
    metric_start_col = len(fixed_headers) + 1

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1
        title_cell = ws.cell(row=1, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.fill = HEADER_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            sub_cell.font = Font(bold=True)
            sub_cell.fill = HEADER_FILL
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        is_last = record["language"] == LANG_ORDER[-1]
        lang_cell = ws.cell(row=row_offset, column=1, value=record["language"])
        lang_cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_last:
            lang_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{m}_{column}"] for m in MODE_ORDER]
            values.append(record[f"best_{column}"])
            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if offset == cols_per_metric - 1 and isinstance(value, str):
                    fill_font = label_fill(value)
                    if fill_font:
                        cell.fill = fill_font[0]
                        cell.font = fill_font[1]
                if is_last:
                    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

    autofit_columns(ws)


def write_oracle_diagnostics_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "oracle_diagnostics"
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, record in enumerate(df.to_dict(orient="records"), start=2):
        for col_idx, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=record[header])
            cell.alignment = Alignment(horizontal="center", vertical="center")

    autofit_columns(ws)


def write_styled_excel(modes_df: pd.DataFrame, diag_df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    write_modes_sheet(wb.active, modes_df)
    write_oracle_diagnostics_sheet(wb.create_sheet(), diag_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--oracle-summary",
        type=Path,
        default=PROJECT_ROOT / "shared" / "results" / "dev_v1" / "original" / "zero_shot" / "gpt" / "metrics_summary.json",
    )
    parser.add_argument(
        "--pipeline-summary",
        type=Path,
        default=EXPERIMENT_DIR / "results" / "metrics_summary.json",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=EXPERIMENT_DIR / "report" / "dev_v1_original_gpt_pipeline_mode_comparison.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if not args.oracle_summary.is_file():
        raise FileNotFoundError(f"Missing: {args.oracle_summary}")
    if not args.pipeline_summary.is_file():
        raise FileNotFoundError(
            f"Missing: {args.pipeline_summary}\n"
            "Run compute_gpt_proposed_metrics.py first."
        )

    oracle_summary = load_summary(args.oracle_summary)
    pipeline_summary = load_summary(args.pipeline_summary)

    modes_df = build_comparison(oracle_summary, pipeline_summary)
    diag_df = build_oracle_diagnostics_table(pipeline_summary)

    output_path = args.output.resolve()
    write_styled_excel(modes_df, diag_df, output_path)
    print(f"Wrote {len(modes_df)} language rows to {output_path}")


if __name__ == "__main__":
    main()
