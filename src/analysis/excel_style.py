"""Shared openpyxl styling for report comparison workbooks.

One rule for every comparison width (2-way, 3-way, N-way): the best value(s) in
a comparison get Good (green), the worst get Bad (red), a genuine tie across
all compared values gets Neutral (yellow), and — for 3+-way comparisons — any
value that is neither best nor worst is left unfilled rather than overloading
"yellow" with a second, unrelated meaning. Colors are Excel's own built-in
"Good/Bad/Neutral" cell-style values, so they read as familiar the moment the
file is opened in Excel.
"""

from __future__ import annotations

from typing import Hashable, TypeVar

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

K = TypeVar("K", bound=Hashable)

GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GOOD_FONT = Font(color="006100")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BAD_FONT = Font(color="9C0006")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NEUTRAL_FONT = Font(color="9C6500")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

THIN = Side(style="thin", color="000000")
THICK = Side(style="medium", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def rank_fills(values: dict[K, float | None]) -> dict[K, tuple[PatternFill, Font]]:
    """Classify each series' value against the others in the same comparison group.

    - Series with a ``None`` value are omitted from the result (no fill).
    - If every present value is equal, every series gets Neutral (tie).
    - Otherwise the max value(s) get Good, the min value(s) get Bad, and any
      value strictly between the max and min (only possible with 3+ series)
      is omitted from the result (left unfilled).
    """
    present = {key: value for key, value in values.items() if value is not None}
    if len(present) < 2:
        return {}

    max_val = max(present.values())
    min_val = min(present.values())
    if max_val == min_val:
        return {key: (NEUTRAL_FILL, NEUTRAL_FONT) for key in present}

    result: dict[K, tuple[PatternFill, Font]] = {}
    for key, value in present.items():
        if value == max_val:
            result[key] = (GOOD_FILL, GOOD_FONT)
        elif value == min_val:
            result[key] = (BAD_FILL, BAD_FONT)
    return result


def label_fill(label: str | None) -> tuple[PatternFill, Font] | None:
    """Fill/font for a ``best_<metric>`` label cell holding a winner name, ``"tie"``, or ``None``."""
    if label is None:
        return None
    if label == "tie":
        return NEUTRAL_FILL, NEUTRAL_FONT
    return GOOD_FILL, GOOD_FONT


def apply_cell_style(
    cell: Cell,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    bold: bool = False,
    border: Border | None = THIN_BORDER,
    thick_bottom: bool = False,
) -> None:
    cell.font = font if font is not None else Font(bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    if border is not None:
        if thick_bottom:
            cell.border = Border(left=border.left, right=border.right, top=border.top, bottom=THICK)
        else:
            cell.border = border


def autofit_columns(ws: Worksheet, *, min_width: int = 8, max_width: int = 40, padding: int = 2) -> None:
    """Size each column from its own cell contents (openpyxl has no true AutoFit)."""
    excluded = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col > merged_range.min_col:
            excluded.add((merged_range.min_row, merged_range.min_col))

    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in excluded:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(width + padding, max_width))
