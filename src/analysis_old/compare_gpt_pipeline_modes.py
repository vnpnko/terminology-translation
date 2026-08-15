"""Compare GPT baseline modes vs gpt_proposed_term pipeline on dev_v1/original.

    python src/analysis/compare_gpt_pipeline_modes.py
    python src/analysis/compare_gpt_pipeline_modes.py \\
        --baseline-dir results/dev_v1/original/gpt \\
        --pipeline-dir results/dev_v1/original/gpt_pipeline
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LANG_ORDER = ("ende", "enru", "enes")
BASELINE_MODES = ("no_term", "proper_term", "random_term")
PIPELINE_MODE = "gpt_proposed_term"
ALL_MODES = BASELINE_MODES + (PIPELINE_MODE,)

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
    (
        "extraction_overlap_pct",
        ("oracle_diagnostics", "extraction_overlap_pct"),
        "Extraction Overlap %",
    ),
    (
        "proposal_match_pct",
        ("oracle_diagnostics", "proposal_match_pct"),
        "Proposal Match %",
    ),
)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BEST_FILLS = {
    "no_term": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "proper_term": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "random_term": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    PIPELINE_MODE: PatternFill(start_color="E2BFED", end_color="E2BFED", fill_type="solid"),
    "tie": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_summary(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict, spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_lang_mode_metrics(summary: dict, mode: str) -> dict[str, dict[str, float | None]]:
    by_lang: dict[str, dict[str, float | None]] = {}
    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue
        mode_data = lang_data.get("modes", {}).get(mode)
        if not mode_data:
            continue
        metrics = mode_data.get("metrics", {})
        by_lang[lang] = {
            column: extract_metric(metrics, spec) for column, spec, _ in METRICS
        }
    return by_lang


def best_mode(values: dict[str, float | None], *, modes: tuple[str, ...]) -> str | None:
    present = {label: value for label, value in values.items() if value is not None}
    if not present:
        return None
    max_val = max(present.values())
    winners = [label for label, value in present.items() if abs(value - max_val) < 1e-9]
    if len(winners) > 1:
        return "tie"
    return winners[0]


def build_comparison(baseline_dir: Path, pipeline_dir: Path) -> pd.DataFrame:
    baseline = load_summary(baseline_dir / "metrics_summary.json")
    pipeline = load_summary(pipeline_dir / "metrics_summary.json")

    baseline_by_mode = {
        mode: extract_lang_mode_metrics(baseline, mode) for mode in BASELINE_MODES
    }
    pipeline_metrics = extract_lang_mode_metrics(pipeline, PIPELINE_MODE)

    rows: list[dict[str, object]] = []
    for lang in LANG_ORDER:
        mode_metrics = {mode: baseline_by_mode[mode].get(lang) for mode in BASELINE_MODES}
        mode_metrics[PIPELINE_MODE] = pipeline_metrics.get(lang)
        if all(metrics is None for metrics in mode_metrics.values()):
            continue

        row: dict[str, object] = {"language": lang}
        compare_modes = BASELINE_MODES + (PIPELINE_MODE,)
        for column, _, _ in METRICS:
            labeled_values: dict[str, float | None] = {}
            for mode in compare_modes:
                metrics = mode_metrics[mode]
                value = metrics.get(column) if metrics else None
                row[f"{mode}_{column}"] = value
                if column not in ("extraction_overlap_pct", "proposal_match_pct"):
                    labeled_values[mode] = value
            row[f"best_{column}"] = best_mode(labeled_values, modes=compare_modes)
        rows.append(row)

    columns = ["language"]
    for column, _, _ in METRICS:
        columns.extend([f"{mode}_{column}" for mode in ALL_MODES])
        columns.append(f"best_{column}")
    return pd.DataFrame(rows, columns=columns)


def write_styled_excel(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "gpt_pipeline_modes"

    fixed_headers = ("language",)
    value_subheaders = ALL_MODES + ("best",)
    cols_per_metric = len(value_subheaders)
    metric_start_col = len(fixed_headers) + 1

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1
        title_cell = ws.cell(row=1, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.fill = HEADER_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = THIN_BORDER
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            sub_cell.font = Font(bold=True)
            sub_cell.fill = HEADER_FILL
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
            sub_cell.border = THIN_BORDER

    for row_offset, record in enumerate(df.to_dict(orient="records"), start=3):
        ws.cell(row=row_offset, column=1, value=record["language"])
        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{mode}_{column}"] for mode in ALL_MODES]
            values.append(record[f"best_{column}"])
            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill = (
                    BEST_FILLS.get(value)
                    if offset == cols_per_metric - 1 and isinstance(value, str)
                    else None
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER
                if fill is not None:
                    cell.fill = fill

    ws.column_dimensions["A"].width = 12
    for col_idx in range(metric_start_col, metric_start_col + len(METRICS) * cols_per_metric):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--baseline-dir",
        type=Path,
        default=Path("results/dev_v1/original/gpt"),
    )
    parser.add_argument(
        "--pipeline-dir",
        type=Path,
        default=Path("results/dev_v1/original/gpt_pipeline"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("report/modes/dev_v1_original_gpt_pipeline_mode_comparison.xlsx"),
    )
    args = parser.parse_args()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    df = build_comparison(args.baseline_dir.resolve(), args.pipeline_dir.resolve())
    write_styled_excel(df, args.output.resolve())
    print(f"Wrote {len(df)} rows to {args.output}")


if __name__ == "__main__":
    main()
