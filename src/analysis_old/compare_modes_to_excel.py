"""Compare no_term, proper_term, and random_term modes across all baseline models.

Writes one Excel file with 9 data rows: 3 models (GPT, Qwen 3B, Qwen 7B),
each with 3 language rows (ende, enru, enes). The model column is merged per block.

    python src/analysis/compare_modes_to_excel.py results/dev_v1/original
    python src/analysis/compare_modes_to_excel.py results/dev_v1/expand
    python src/analysis/compare_modes_to_excel.py results/dev_v1/cleaned
    python src/analysis/compare_modes_to_excel.py results/dev_v2
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LANG_ORDER = ("ende", "enru", "enes")
MODE_ORDER = ("no_term", "proper_term", "random_term")
BASELINE_DIRS = ("gpt", "qwen_3b", "qwen_7b")
BASELINE_LABELS = {
    "gpt": "GPT",
    "qwen_3b": "Qwen 3B",
    "qwen_7b": "Qwen 7B",
}

METRICS = (
    ("bleu", "bleu", "BLEU"),
    ("chrf", "chrf", "chrF"),
    ("term_accuracy_pct", ("terminology_accuracy", "avg_ratio_pct"), "Term Accuracy %"),
    ("macro_avg_consistency", ("terminology_consistency", "macro_avg_consistency"), "Macro Consistency"),
    (
        "weighted_avg_consistency",
        ("terminology_consistency", "weighted_avg_consistency"),
        "Weighted Consistency",
    ),
)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BEST_FILLS = {
    "no_term": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "proper_term": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "random_term": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "tie": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

THIN = Side(style="thin", color="000000")
THICK = Side(style="medium", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_summary(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def extract_metric(metrics: dict, spec: str | tuple[str, str]) -> float | None:
    if isinstance(spec, str):
        value = metrics.get(spec)
    else:
        section, key = spec
        value = metrics.get(section, {}).get(key)
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def extract_mode_metrics(summary: dict) -> dict[tuple[str, str], dict[str, float | None]]:
    by_lang_mode: dict[tuple[str, str], dict[str, float | None]] = {}

    for lang in LANG_ORDER:
        lang_data = summary.get("languages", {}).get(lang)
        if not lang_data:
            continue

        for mode in MODE_ORDER:
            mode_data = lang_data.get("modes", {}).get(mode)
            if not mode_data:
                continue

            metrics = mode_data.get("metrics", {})
            by_lang_mode[(lang, mode)] = {
                column: extract_metric(metrics, spec) for column, spec, _ in METRICS
            }

    return by_lang_mode


def best_mode(values: dict[str, float | None]) -> str | None:
    present = {label: value for label, value in values.items() if value is not None}
    if not present:
        return None

    max_val = max(present.values())
    winners = [label for label, value in present.items() if abs(value - max_val) < 1e-9]
    if len(winners) > 1:
        return "tie"
    return winners[0]


def build_comparison(dataset_dir: Path) -> pd.DataFrame:
    summaries: dict[str, dict[tuple[str, str], dict[str, float | None]]] = {}

    for baseline_dir in BASELINE_DIRS:
        summary_path = dataset_dir / baseline_dir / "metrics_summary.json"
        if not summary_path.exists():
            raise FileNotFoundError(f"Missing metrics file: {summary_path}")
        summaries[baseline_dir] = extract_mode_metrics(load_summary(summary_path))

    rows: list[dict[str, object]] = []

    for baseline_dir in BASELINE_DIRS:
        for lang in LANG_ORDER:
            mode_metrics = {
                mode: summaries[baseline_dir].get((lang, mode)) for mode in MODE_ORDER
            }
            if all(metrics is None for metrics in mode_metrics.values()):
                continue

            row: dict[str, object] = {
                "model": BASELINE_LABELS[baseline_dir],
                "language": lang,
            }

            for column, _, _ in METRICS:
                labeled_values: dict[str, float | None] = {}
                for mode in MODE_ORDER:
                    metrics = mode_metrics[mode]
                    value = metrics.get(column) if metrics else None
                    row[f"{mode}_{column}"] = value
                    labeled_values[mode] = value

                row[f"best_{column}"] = best_mode(labeled_values)

            rows.append(row)

    columns = ["model", "language"]
    for column, _, _ in METRICS:
        columns.extend([f"{mode}_{column}" for mode in MODE_ORDER])
        columns.append(f"best_{column}")

    return pd.DataFrame(rows, columns=columns)


def apply_cell_style(
    cell,
    *,
    bold: bool = False,
    fill: PatternFill | None = None,
    border: Border | None = THIN_BORDER,
    thick_bottom: bool = False,
) -> None:
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    if border is not None:
        if thick_bottom:
            cell.border = Border(
                left=border.left,
                right=border.right,
                top=border.top,
                bottom=THICK,
            )
        else:
            cell.border = border


def write_styled_excel(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "modes"

    fixed_headers = ("model", "language")
    value_subheaders = MODE_ORDER + ("best",)
    cols_per_metric = len(value_subheaders)

    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    metric_start_col = len(fixed_headers) + 1
    for metric_idx, (_, _, title) in enumerate(METRICS):
        start_col = metric_start_col + metric_idx * cols_per_metric
        end_col = start_col + cols_per_metric - 1

        title_cell = ws.cell(row=1, column=start_col, value=title)
        apply_cell_style(title_cell, bold=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, subheader in enumerate(value_subheaders):
            sub_cell = ws.cell(row=2, column=start_col + offset, value=subheader)
            apply_cell_style(sub_cell, bold=True, fill=HEADER_FILL)

    records = df.to_dict(orient="records")
    for row_offset, record in enumerate(records, start=3):
        thick_bottom = record["language"] == "enes"

        model_cell = ws.cell(
            row=row_offset,
            column=1,
            value=record["model"] if record["language"] == LANG_ORDER[0] else None,
        )
        apply_cell_style(model_cell, thick_bottom=thick_bottom)

        lang_cell = ws.cell(row=row_offset, column=2, value=record["language"])
        apply_cell_style(lang_cell, thick_bottom=thick_bottom)

        for metric_idx, (column, _, _) in enumerate(METRICS):
            start_col = metric_start_col + metric_idx * cols_per_metric
            values = [record[f"{mode}_{column}"] for mode in MODE_ORDER]
            values.append(record[f"best_{column}"])

            for offset, value in enumerate(values):
                cell = ws.cell(row=row_offset, column=start_col + offset, value=value)
                fill = (
                    BEST_FILLS.get(value)
                    if offset == cols_per_metric - 1 and isinstance(value, str)
                    else None
                )
                apply_cell_style(cell, fill=fill, thick_bottom=thick_bottom)

    for baseline_idx in range(len(BASELINE_DIRS)):
        start_row = 3 + baseline_idx * len(LANG_ORDER)
        end_row = start_row + len(LANG_ORDER) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    for col_idx in range(metric_start_col, metric_start_col + len(METRICS) * cols_per_metric):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(output_path)


def dataset_slug(dataset_dir: Path) -> str:
    parts = list(dataset_dir.resolve().parts)
    if "results" in parts:
        parts = parts[parts.index("results") + 1 :]
    return "_".join(parts)


def default_output_path(dataset_dir: Path, report_dir: Path) -> Path:
    return report_dir / "modes" / f"{dataset_slug(dataset_dir)}_mode_comparison.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "dataset_dir",
        type=Path,
        help="Results directory containing gpt/, qwen_3b/, and qwen_7b/ subfolders",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: report/modes/<dataset>_mode_comparison.xlsx)",
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=Path("report"),
        help="Report output directory when --output is not set",
    )
    args = parser.parse_args()

    dataset_dir = args.dataset_dir.resolve()
    output_path = (
        args.output.resolve()
        if args.output
        else default_output_path(dataset_dir, args.report_dir.resolve())
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = build_comparison(dataset_dir)
    write_styled_excel(df, output_path)

    print(
        f"Wrote {len(df)} rows "
        f"({len(BASELINE_DIRS)} models x {len(LANG_ORDER)} languages) to {output_path}"
    )


if __name__ == "__main__":
    main()
